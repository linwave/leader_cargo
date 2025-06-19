import datetime
import json
import os
import uuid
from concurrent.futures import ThreadPoolExecutor

from PIL import Image
from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db.models import F, Q, FloatField, Sum, Count
from django.db.models.functions import Cast, TruncMonth
from django.http import HttpResponse, Http404, FileResponse, JsonResponse
from django.utils.decorators import method_decorator
from django.utils.safestring import mark_safe
from django.utils.timezone import make_aware
from django.shortcuts import redirect, render
from django.urls import reverse_lazy, reverse
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import CreateView, DeleteView, UpdateView, TemplateView, ListView

from telegram_bot.models import TelegramProfile
from telegram_bot.utils import send_telegram_message
from .forms import AddCarrierFilesForm, EditTableArticleForm, EditTransportationTariffForClients, AddCarriersListForm, EditCarriersListForm, DeleteCarriersListForm, AddRoadForm, EditRoadForm, DeleteRoadForm, AddRoadToCarriersForm, \
    DeleteRoadToCarriersForm, AddRequestsForLogisticsCalculationsForm, EditRequestsForLogisticsCalculationsForm, AddGoodsRequestLogisticsForm, NewStatusRequestForm, EditGoodsRequestLogisticsForm, EditRoadToCarriersForm, \
    EditPaidByTheClientArticleForm, AddBidRequestLogisticsForm, UpdateRequestForm, AddCargo
from .models import CargoFiles, CargoArticle, RequestsForLogisticsCalculations, CarriersList, RoadsList, RequestsForLogisticsGoods, RequestsForLogisticFiles, CarriersRoadParameters, PriceListsOfCarriers, PaymentDocumentsForArticles, \
    RequestsForLogisticsRate
from django.shortcuts import get_object_or_404
# FROM MAIN
from main.models import CustomUser
from main.utils import DataMixin, MyLoginMixin
# EXTERNAL LIBRARIES
import openpyxl
import xlrd
import xlsxwriter as xlsxwriter


def download(request, file_id):
    obj = PaymentDocumentsForArticles.objects.get(pk=file_id)
    if obj:
        return FileResponse(obj.file_path)
    raise Http404


class LogisticRequestsView(MyLoginMixin, DataMixin, TemplateView):
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП', 'Менеджер']
    template_name = 'analytics/logistic_requests/logistics_requests.html'
    login_url = reverse_lazy('main:login')
    message = dict()

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.role == 'Супер Администратор' or self.request.user.role == 'Логист' or self.request.user.role == 'РОП':
            context['reports'] = RequestsForLogisticsCalculations.objects.all().select_related('initiator')
        else:
            context['reports'] = RequestsForLogisticsCalculations.objects.filter(initiator=self.request.user.pk).select_related('initiator')

        if self.request.user.role == 'РОП':
            context['reports_work'] = context['reports'].exclude(status='Черновик').filter(initiator__town=self.request.user.town)
        else:
            context['reports_work'] = context['reports'].exclude(status='Черновик')

        if self.request.user.role != 'Логист':
            if self.request.user.role == 'РОП':
                context['reports_draft'] = context['reports'].filter(status='Черновик').filter(initiator=self.request.user.pk)
            else:
                context['reports_draft'] = context['reports'].filter(status='Черновик')

        context['table_paginator'] = Paginator(context['reports_work'], 100)
        page_number = self.request.GET.get('page', 1)
        context['table_paginator_obj'] = context['table_paginator'].get_page(page_number)

        context['reports'] = context['table_paginator_obj']

        if self.request.user.role == 'Логист':
            c_def = self.get_user_context(title="Обработка запросов")
        else:
            c_def = self.get_user_context(title="Запрос ставки")

        return dict(list(context.items()) + list(c_def.items()))


def requests_create_all_files(request):
    template_name = 'analytics/logistic_requests/partial/create_file_for_logist.html'
    error = False
    message_error = None
    path_for_file_template = None
    reports_name = []
    if request.POST:
        if request.POST.get('csrfmiddlewaretoken'):
            try:
                filename = "%s.%s" % (uuid.uuid4(), '.xlsx')
                path_for_file = os.path.join(f'/media/files/logistic/requests/cash/', filename)
                if settings.DEBUG:
                    name_for_excel = str(os.getcwd()) + path_for_file
                    os.makedirs(f'media/files/logistic/requests/cash/', exist_ok=True)
                else:
                    name_for_excel = str(os.getcwd()) + '/leader_cargo' + path_for_file
                    os.makedirs(f'leader_cargo/media/files/logistic/requests/cash/', exist_ok=True)
                workbook = xlsxwriter.Workbook(f'{name_for_excel}')
                worksheet = workbook.add_worksheet(name='Packing list')
                worksheet.set_row(0, 60)
                worksheet.set_column(0, 0, 60)
                worksheet.set_column(1, 2, 40)
                worksheet.set_column(3, 10, 33)
                titles = workbook.add_format({'bold': True, 'bg_color': '#FFFACD', 'border': 1})
                titles.set_align('center')
                titles.set_align('vcenter')
                titles.set_font_size(14)
                titles.set_font_name('Times New Roman')
                titles.set_text_wrap()
                worksheet.write('A1', 'фото\n图片', titles)
                worksheet.write('B1', 'описание\n品名', titles)
                worksheet.write('C1', 'материал\n材', titles)
                worksheet.write('D1', 'количество упаковок/мест\n箱数', titles)
                worksheet.write('E1', 'количество в каждой упаковке (шт)\n每箱里面的数量', titles)
                worksheet.write('F1', 'объём/размер упаковки (м3)\n箱子尺寸', titles)
                worksheet.write('G1', 'вес брутто упаковки (кг)\n重量', titles)
                worksheet.write('H1', 'общий объём (м3)\n总体积 ', titles)
                worksheet.write('I1', 'общий вес брутто (кг)\n总重量', titles)
                worksheet.write('J1', 'общее кол-во (шт)\n数量', titles)
                worksheet.write('K1', 'торговая марка\n商标', titles)
                cell_format = workbook.add_format()
                cell_format.set_align('center')
                cell_format.set_align('vcenter')
                cell_format.set_font_size(12)
                cell_format.set_font_name('Times New Roman')
                cell_format.set_border()
                cell_format.set_text_wrap()
                count = 2
                for report in request.POST:
                    if "report" in report:
                        new_report = RequestsForLogisticsCalculations.objects.get(pk=report.split("_")[2])
                        reports_name.append(f'{new_report.name}')
                        all_goods = new_report.goods.all()
                        for row, good in enumerate(all_goods, start=count):
                            if good.photo_path_logistic_goods:
                                if settings.DEBUG:
                                    path_photo = str(os.getcwd()) + good.photo_path_logistic_goods.url
                                else:
                                    path_photo = str(os.getcwd()) + '/leader_cargo' + good.photo_path_logistic_goods.url
                                # worksheet.insert_image(f'A{row}', path_photo, {'object_position': 1, 'x_offset': 20, 'y_offset': 5})
                                worksheet.write(f'A{row}', '', cell_format)
                                worksheet.insert_image(f'A{row}', path_photo, {'object_position': 1})
                            worksheet.set_row(row - 1, 318)
                            if good.description:
                                worksheet.write(f'B{row}', good.description.strip(), cell_format)
                            else:
                                worksheet.write(f'B{row}', '', cell_format)
                            if good.material:
                                worksheet.write(f'C{row}', good.material.strip(), cell_format)
                            else:
                                worksheet.write(f'C{row}', '', cell_format)
                            worksheet.write(f'D{row}', good.number_of_packages, cell_format)
                            worksheet.write(f'E{row}', good.quantity_in_each_package, cell_format)
                            worksheet.write(f'F{row}', good.size_of_packaging, cell_format)
                            worksheet.write(f'G{row}', good.gross_weight_of_packaging, cell_format)
                            worksheet.write(f'H{row}', good.total_volume, cell_format)
                            worksheet.write(f'I{row}', good.total_gross_weight, cell_format)
                            worksheet.write(f'J{row}', good.total_quantity, cell_format)
                            worksheet.write(f'K{row}', good.trademark, cell_format)
                        count = count + len(all_goods)
                worksheet.autofit()
                workbook.close()
                path_for_file_template = path_for_file
            except Exception as e:
                error = True
                message_error = e
            return render(request, template_name,
                          context={"name_for_excel": path_for_file_template,
                                   "reports_name": reports_name,
                                   "error": error,
                                   "message_error": message_error})
    else:
        return render(request, template_name,
                      context={})


def LogisticRequestsViewAutoUpdate(request):
    template_name = 'analytics/logistic_requests/partial/logistic_requests_auto_update.html'
    reports = RequestsForLogisticsCalculations.objects.exclude(status='Черновик')
    return render(request, template_name, {'reports_work': reports})


class LogisticRequestsAddView(MyLoginMixin, DataMixin, CreateView):
    model = RequestsForLogisticsCalculations
    form_class = AddRequestsForLogisticsCalculationsForm
    template_name = 'analytics/logistic_requests/logistics_requests_add.html'
    login_url = reverse_lazy('main:login')
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП', 'Менеджер']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Создание запроса")
        return dict(list(context.items()) + list(c_def.items()))

    def form_valid(self, form):
        new_data = form.save(commit=False)
        new_data.initiator = self.request.user
        new_data.status = 'Черновик'
        new_data.save()
        return redirect('analytics:edit_logistic_requests', new_data.pk)


class LogisticRequestsEditView(MyLoginMixin, DataMixin, UpdateView):
    model = RequestsForLogisticsCalculations
    form_class = EditRequestsForLogisticsCalculationsForm
    pk_url_kwarg = 'request_id'
    template_name = 'analytics/logistic_requests/logistics_requests_edit.html'
    login_url = reverse_lazy('main:login')
    success_url = reverse_lazy('analytics:logistic_requests')
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП', 'Менеджер']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['my_request'] = self.get_object()
        context['roads'] = RoadsList.objects.all()
        context['all_carriers'] = CarriersList.objects.all()
        context['my_request_description_roads'] = context['my_request'].roads.all().values_list('name', flat=True)
        if context['my_request'].notification and self.request.user.pk == context['my_request'].initiator.pk:
            context['my_request'].notification = False
            context['my_request'].save()
        # context['all_bids'] = context['my_request'].rate.all().order_by("road")
        context['all_documents'] = context['my_request'].requestsforlogisticfiles_set.all()
        context['goods'] = context['my_request'].goods.all()
        c_def = self.get_user_context(title="Редактирование запроса на просчет")
        return dict(list(context.items()) + list(c_def.items()))

    def form_valid(self, form):
        new_data = form.save(commit=False)
        if new_data.status == 'Черновик':
            if 'status_new' in self.request.POST:
                new_data.status = 'Новый'
                new_data.notification = False
                new_data.save()
            for road in new_data.roads.all():
                new_data.roads.remove(road)
            for req in self.request.POST:
                if 'road' in req:
                    road_pk = req.split('-')[1]
                    road = RoadsList.objects.get(pk=road_pk)
                    new_data.roads.add(road)
                if 'comments_initiator' in req:
                    new_data.comments_initiator = self.request.POST.get('comments_initiator')
            for f in self.request.FILES.getlist('files_for_request'):
                new_data.requestsforlogisticfiles_set.create(name=f, file_path_request=f)
            new_data.save()
            if new_data.goods.all():
                if not new_data.file_path_request:
                    filename = "%s.%s" % (uuid.uuid4(), '.xlsx')
                    new_data.file_path_request = os.path.join(f'files/logistic/requests/{datetime.datetime.now().strftime("%Y/%m/%d/")}', filename)
                    new_data.save()
                if settings.DEBUG:
                    name_for_excel = str(os.getcwd()) + new_data.file_path_request.url
                    os.makedirs(f'media/files/logistic/requests/{datetime.datetime.now().strftime("%Y/%m/%d/")}', exist_ok=True)
                else:
                    name_for_excel = str(os.getcwd()) + '/leader_cargo' + new_data.file_path_request.url
                    os.makedirs(f'leader_cargo/media/files/logistic/requests/{datetime.datetime.now().strftime("%Y/%m/%d/")}', exist_ok=True)
                workbook = xlsxwriter.Workbook(f'{name_for_excel}')
                worksheet = workbook.add_worksheet(name='Packing list')
                worksheet.set_row(0, 60)
                worksheet.set_column(0, 0, 60)
                worksheet.set_column(1, 2, 40)
                worksheet.set_column(3, 10, 33)
                titles = workbook.add_format({'bold': True, 'bg_color': '#FFFACD', 'border': 1})
                titles.set_align('center')
                titles.set_align('vcenter')
                titles.set_font_size(14)
                titles.set_font_name('Times New Roman')
                titles.set_text_wrap()
                worksheet.write('A1', 'фото\n图片', titles)
                worksheet.write('B1', 'описание\n品名', titles)
                worksheet.write('C1', 'материал\n材', titles)
                worksheet.write('D1', 'количество упаковок/мест\n箱数', titles)
                worksheet.write('E1', 'количество в каждой упаковке (шт)\n每箱里面的数量', titles)
                worksheet.write('F1', 'объём/размер упаковки (м3)\n箱子尺寸', titles)
                worksheet.write('G1', 'вес брутто упаковки (кг)\n重量', titles)
                worksheet.write('H1', 'общий объём (м3)\n总体积 ', titles)
                worksheet.write('I1', 'общий вес брутто (кг)\n总重量', titles)
                worksheet.write('J1', 'общее кол-во (шт)\n数量', titles)
                worksheet.write('K1', 'торговая марка\n商标', titles)
                all_goods = new_data.goods.all()
                cell_format = workbook.add_format()
                cell_format.set_align('center')
                cell_format.set_align('vcenter')
                cell_format.set_font_size(12)
                cell_format.set_font_name('Times New Roman')
                cell_format.set_border()
                cell_format.set_text_wrap()
                for row, good in enumerate(all_goods, start=2):
                    if good.photo_path_logistic_goods:
                        if settings.DEBUG:
                            path_photo = str(os.getcwd()) + good.photo_path_logistic_goods.url
                        else:
                            path_photo = str(os.getcwd()) + '/leader_cargo' + good.photo_path_logistic_goods.url
                        # worksheet.insert_image(f'A{row}', path_photo, {'object_position': 1, 'x_offset': 20, 'y_offset': 5})
                        worksheet.write(f'A{row}', '', cell_format)
                        worksheet.insert_image(f'A{row}', path_photo, {'object_position': 1})
                    worksheet.set_row(row - 1, 318)
                    if good.description:
                        worksheet.write(f'B{row}', good.description.strip(), cell_format)
                    else:
                        worksheet.write(f'B{row}', '', cell_format)
                    if good.material:
                        worksheet.write(f'C{row}', good.material.strip(), cell_format)
                    else:
                        worksheet.write(f'C{row}', '', cell_format)
                    worksheet.write(f'D{row}', good.number_of_packages, cell_format)
                    worksheet.write(f'E{row}', good.quantity_in_each_package, cell_format)
                    worksheet.write(f'F{row}', good.size_of_packaging, cell_format)
                    worksheet.write(f'G{row}', good.gross_weight_of_packaging, cell_format)
                    worksheet.write(f'H{row}', good.total_volume, cell_format)
                    worksheet.write(f'I{row}', good.total_gross_weight, cell_format)
                    worksheet.write(f'J{row}', good.total_quantity, cell_format)
                    worksheet.write(f'K{row}', good.trademark, cell_format)
                worksheet.autofit()
                workbook.close()
        if new_data.status == 'Новый':
            logists_china = CustomUser.objects.filter(role='Логист Китай')
            for log in logists_china:
                telegram_profile, created = TelegramProfile.objects.get_or_create(user=log)
                if telegram_profile.is_verified:
                    send_telegram_message(telegram_profile.chat_id, f"Новый запрос {new_data}\n"
                                                                    f"https://nextkargo.ru/logistic/logistic-requests/edit/{new_data.pk}", settings.TELEGRAM_BOT_TOKEN)
        return redirect('analytics:edit_logistic_requests', new_data.pk)

    def get(self, request, *args, **kwargs):
        if self.get_object().status == 'Черновик' and self.request.user.role == 'Логист':
            return redirect('analytics:logistic_requests')
        return super(LogisticRequestsEditView, self).get(request, *args, **kwargs)


def editLogisticRequest(request, request_id):
    my_request = RequestsForLogisticsCalculations.objects.get(pk=request_id)
    # if request.FILES:
    #     for file in request.FILES["files_for_request"]:
    #         my_request.requestsforlogisticfiles_set.create(name=file, file_path_request=file)
    if request.POST:
        if 'comments_logist' in request.POST:
            my_request.comments_logist = request.POST['comments_logist']
            my_request.notification = True
        if 'tariff_for_client' in request.POST:
            my_request.tariff_for_client = request.POST['tariff_for_client']
        my_request.save()
    return JsonResponse({'success': True})


class NewStatusRequest(MyLoginMixin, DataMixin, UpdateView):
    model = RequestsForLogisticsCalculations
    form_class = NewStatusRequestForm
    pk_url_kwarg = 'request_id'
    context_object_name = 'my_request'
    template_name = 'analytics/logistic_requests/modal/request_status_new.html'
    login_url = reverse_lazy('main:login')
    success_url = reverse_lazy('analytics:logistic_requests')
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП', 'Менеджер']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Отправление запроса логисту")
        return dict(list(context.items()) + list(c_def.items()))

    def form_valid(self, form):
        new_data = form.save(commit=False)
        if new_data.status == 'Черновик':
            new_data.status = 'Новый'
            new_data.notification = False
            new_data.save()
            for req in self.request.POST:
                if 'road' in req:
                    road_pk = req.split('-')[1]
                    road = RoadsList.objects.get(pk=road_pk)
                    new_data.roads.add(road)
                if 'comments_initiator' in req:
                    new_data.comments_initiator = self.request.POST.get('comments_initiator')
            for f in self.request.FILES.getlist('files_for_request'):
                new_data.requestsforlogisticfiles_set.create(name=f, file_path_request=f)
            if new_data.goods.all():
                if not new_data.file_path_request:
                    filename = "%s.%s" % (uuid.uuid4(), '.xlsx')
                    new_data.file_path_request = os.path.join(f'files/logistic/requests/{datetime.datetime.now().strftime("%Y/%m/%d/")}', filename)
                new_data.save()
                if settings.DEBUG:
                    name_for_excel = str(os.getcwd()) + new_data.file_path_request.url
                    os.makedirs(f'media/files/logistic/requests/{datetime.datetime.now().strftime("%Y/%m/%d/")}', exist_ok=True)
                else:
                    name_for_excel = str(os.getcwd()) + '/leader_cargo' + new_data.file_path_request.url
                    os.makedirs(f'leader_cargo/media/files/logistic/requests/{datetime.datetime.now().strftime("%Y/%m/%d/")}', exist_ok=True)
                workbook = xlsxwriter.Workbook(f'{name_for_excel}')
                worksheet = workbook.add_worksheet(name='Packing list')
                worksheet.set_row(0, 60)
                worksheet.set_column(0, 0, 60)
                worksheet.set_column(1, 2, 40)
                worksheet.set_column(3, 10, 33)
                titles = workbook.add_format({'bold': True, 'bg_color': '#FFFACD', 'border': 1})
                titles.set_align('center')
                titles.set_align('vcenter')
                titles.set_font_size(14)
                titles.set_font_name('Times New Roman')
                titles.set_text_wrap()
                worksheet.write('A1', 'фото\n图片', titles)
                worksheet.write('B1', 'описание\n品名', titles)
                worksheet.write('C1', 'материал\n材', titles)
                worksheet.write('D1', 'количество упаковок/мест\n箱数', titles)
                worksheet.write('E1', 'количество в каждой упаковке (шт)\n每箱里面的数量', titles)
                worksheet.write('F1', 'объём/размер упаковки (м3)\n箱子尺寸', titles)
                worksheet.write('G1', 'вес брутто упаковки (кг)\n重量', titles)
                worksheet.write('H1', 'общий объём (м3)\n总体积 ', titles)
                worksheet.write('I1', 'общий вес брутто (кг)\n总重量', titles)
                worksheet.write('J1', 'общее кол-во (шт)\n数量', titles)
                worksheet.write('K1', 'торговая марка\n商标', titles)
                all_goods = new_data.goods.all()
                cell_format = workbook.add_format()
                cell_format.set_align('center')
                cell_format.set_align('vcenter')
                cell_format.set_font_size(12)
                cell_format.set_font_name('Times New Roman')
                cell_format.set_border()
                cell_format.set_text_wrap()
                for row, good in enumerate(all_goods, start=2):
                    if good.photo_path_logistic_goods:
                        if settings.DEBUG:
                            path_photo = str(os.getcwd()) + good.photo_path_logistic_goods.url
                        else:
                            path_photo = str(os.getcwd()) + '/leader_cargo' + good.photo_path_logistic_goods.url
                        worksheet.write(f'A{row}', '', cell_format)
                        worksheet.insert_image(f'A{row}', path_photo, {'object_position': 1})
                    worksheet.set_row(row - 1, 318)
                    if good.description:
                        worksheet.write(f'B{row}', good.description.strip(), cell_format)
                    else:
                        worksheet.write(f'B{row}', '', cell_format)
                    if good.material:
                        worksheet.write(f'C{row}', good.material.strip(), cell_format)
                    else:
                        worksheet.write(f'C{row}', '', cell_format)
                    worksheet.write(f'D{row}', good.number_of_packages, cell_format)
                    worksheet.write(f'E{row}', good.quantity_in_each_package, cell_format)
                    worksheet.write(f'F{row}', good.size_of_packaging, cell_format)
                    worksheet.write(f'G{row}', good.gross_weight_of_packaging, cell_format)
                    worksheet.write(f'H{row}', good.total_volume, cell_format)
                    worksheet.write(f'I{row}', good.total_gross_weight, cell_format)
                    worksheet.write(f'J{row}', good.total_quantity, cell_format)
                    worksheet.write(f'K{row}', good.trademark, cell_format)
                worksheet.autofit()
                workbook.close()
                return redirect('analytics:edit_logistic_requests', new_data.pk)
            else:
                new_data.save()
            return redirect('analytics:edit_logistic_requests', new_data.pk)
        return redirect('analytics:edit_logistic_requests', new_data.pk)


class LogisticRequestsBackToManagerView(MyLoginMixin, DataMixin, UpdateView):
    model = RequestsForLogisticsCalculations
    form_class = UpdateRequestForm
    pk_url_kwarg = 'request_id'
    context_object_name = 'my_request'
    template_name = 'analytics/logistic_requests/modal/new_to_draft.html'
    login_url = reverse_lazy('main:login')
    success_url = reverse_lazy('analytics:logistic_requests')
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП', 'Менеджер']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Отправление запроса обратно менеджеру")
        return dict(list(context.items()) + list(c_def.items()))

    def form_valid(self, form):
        new_data = form.save(commit=False)
        if new_data.status == 'Новый' or new_data.status == 'В работе' or new_data.status == 'Запрос на изменение' or new_data.status == 'На просчете':
            new_data.status = 'Черновик'
            new_data.notification = True
            new_data.save()
            return redirect('analytics:logistic_requests')
        return redirect('analytics:edit_logistic_requests', new_data.pk)


class LogisticRequestsForEditView(MyLoginMixin, DataMixin, UpdateView):
    model = RequestsForLogisticsCalculations
    form_class = UpdateRequestForm
    pk_url_kwarg = 'request_id'
    context_object_name = 'my_request'
    template_name = 'analytics/logistic_requests/modal/all_to_for_edit.html'
    login_url = reverse_lazy('main:login')
    success_url = reverse_lazy('analytics:logistic_requests')
    role_have_perm = ['Супер Администратор', 'РОП', 'Менеджер']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Замена статуса на Запрос на изменение")
        return dict(list(context.items()) + list(c_def.items()))

    def form_valid(self, form):
        new_data = form.save(commit=False)
        new_data.status = 'Запрос на изменение'
        new_data.save()
        return redirect('analytics:edit_logistic_requests', new_data.pk)


def work_status_request(request, request_id):
    role_have_perm = ['Супер Администратор', 'Логист']
    my_request = RequestsForLogisticsCalculations.objects.get(pk=request_id)
    if request.user.role in role_have_perm:
        if my_request.status == 'Новый':
            my_request.status = 'В работе'
        my_request.save()
    return redirect('analytics:edit_logistic_requests', my_request.pk)


def calculation_status_request(request, request_id):
    role_have_perm = ['Супер Администратор', 'Логист']
    my_request = RequestsForLogisticsCalculations.objects.get(pk=request_id)
    if request.user.role in role_have_perm:
        if my_request.status == 'В работе':
            my_request.status = 'На просчете'
        my_request.save()
    return redirect('analytics:edit_logistic_requests', my_request.pk)


def progress_status_request(request, request_id):
    role_have_perm = ['Супер Администратор', 'Логист']
    my_request = RequestsForLogisticsCalculations.objects.get(pk=request_id)
    if request.user.role in role_have_perm:
        if my_request.status == 'На просчете' or my_request.status == 'Частично обработано':
            my_request.status = 'Обработано'
        my_request.save()
    return redirect('analytics:edit_logistic_requests', my_request.pk)


class LogisticRequestsCloseStatusView(MyLoginMixin, DataMixin, UpdateView):
    model = RequestsForLogisticsCalculations
    form_class = UpdateRequestForm
    pk_url_kwarg = 'request_id'
    context_object_name = 'my_request'
    template_name = 'analytics/logistic_requests/modal/close_request.html'
    login_url = reverse_lazy('main:login')
    success_url = reverse_lazy('analytics:logistic_requests')
    role_have_perm = ['Супер Администратор', 'РОП', 'Менеджер']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['my_request'] = self.get_object()
        context['my_select'] = 0

        context['reasons'] = {
            "1": "Клиент отказался",
            "2": "Слишком долго делали просчет",
            "3": "Клиента не устроила цена выкупа",
            "4": "Дорогая логистика",
            "5": "Другое",
            "6": "Ставка выбрана",
        }
        if self.request.GET:
            context['my_select'] = self.request.GET["reason"]
            if context['my_select'] == '6':
                context['my_goods'] = context['my_request'].goods.all()
                # context['bids'] = context['my_request'].rate.all()
        c_def = self.get_user_context(title="Закрытие запроса")
        return dict(list(context.items()) + list(c_def.items()))

    def form_valid(self, form):
        new_data = form.save(commit=False)
        if new_data.status == 'Частично обработано' or new_data.status == 'Обработано':
            new_data.status = 'Закрыт'
            for req in self.request.POST:
                if 'good-bid' in req:
                    good_pk = req.split("-")[2]
                    bid_pk = self.request.POST[req].split("-")[1]
                    good = RequestsForLogisticsGoods.objects.get(pk=good_pk)
                    bid = good.rate.get(pk=bid_pk)
                    good.bid = bid.bid
                    good.save()
                    new_data.carrier = CarriersList.objects.get(pk=bid.carrier.pk)
        if self.request.POST:
            reasons = {
                "1": "Клиент отказался",
                "2": "Слишком долго делали просчет",
                "3": "Клиента не устроила цена выкупа",
                "4": "Дорогая логистика",
                "5": "Другое",
                "6": "Ставка выбрана",
            }
            for key, value in reasons.items():
                if key == self.request.POST.get('reason', 5):
                    new_data.reason_for_close = value
                    break
        new_data.save()
        return redirect('analytics:edit_logistic_requests', new_data.pk)


class AddGoodsLogisticRequestsView(MyLoginMixin, DataMixin, CreateView):
    model = RequestsForLogisticsGoods
    form_class = AddGoodsRequestLogisticsForm
    pk_url_kwarg = 'request_id'
    template_name = 'analytics/logistic_requests/partial/add_goods_for_requests_logistic.html'
    login_url = reverse_lazy('main:login')
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП', 'Менеджер']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['good'] = RequestsForLogisticsGoods.objects.last()
        context['my_request'] = context['good'].request
        context['count_goods'] = RequestsForLogisticsGoods.objects.filter(request=self.kwargs['request_id']).count()
        c_def = self.get_user_context(title="Создание товара к запросу")
        return dict(list(context.items()) + list(c_def.items()))

    def get(self, request, *args, **kwargs):
        my_request = get_object_or_404(RequestsForLogisticsCalculations, pk=self.kwargs['request_id'])
        my_request.goods.create()
        return super().get(request, *args, **kwargs)


def editGoodsLogisticRequests(request, goods_id):
    good = RequestsForLogisticsGoods.objects.get(pk=goods_id)
    if request.GET.get('delete'):
        good.photo_path_logistic_goods.delete(save=False)
        good.save()
        return render(request, 'analytics/logistic_requests/partial/edit_goods_for_requests_logistic.html',
                      {'good': good})
    if request.FILES:
        good.photo_path_logistic_goods = request.FILES["photo_path_logistic_goods"]
        good.save()
        if settings.DEBUG:
            path_photo = str(os.getcwd()) + good.photo_path_logistic_goods.url
        else:
            path_photo = str(os.getcwd()) + '/leader_cargo' + good.photo_path_logistic_goods.url
        with Image.open(path_photo) as img:
            img.thumbnail(size=(426, 424))
            img.save(path_photo)
    elif request.POST:
        if 'description' in request.POST:
            good.description = request.POST['description']
        elif 'material' in request.POST:
            good.material = request.POST['material']
        elif 'number_of_packages' in request.POST:
            good.number_of_packages = request.POST['number_of_packages']
        elif 'quantity_in_each_package' in request.POST:
            good.quantity_in_each_package = request.POST['quantity_in_each_package']
        elif 'size_of_packaging' in request.POST:
            good.size_of_packaging = request.POST['size_of_packaging']
        elif 'gross_weight_of_packaging' in request.POST:
            good.gross_weight_of_packaging = request.POST['gross_weight_of_packaging']
        elif 'total_volume' in request.POST:
            good.total_volume = request.POST['total_volume']
        elif 'total_gross_weight' in request.POST:
            good.total_gross_weight = request.POST['total_gross_weight']
        elif 'total_quantity' in request.POST:
            good.total_quantity = request.POST['total_quantity']
        elif 'trademark' in request.POST:
            good.trademark = request.POST['trademark']
        good.save()
    return render(request, 'analytics/logistic_requests/partial/edit_goods_for_requests_logistic.html',
                  {'good': good})


class DeleteGoodsLogisticRequestsView(MyLoginMixin, DataMixin, DeleteView):
    model = RequestsForLogisticsGoods
    pk_url_kwarg = 'goods_id'
    template_name = 'analytics/logistic_requests/partial/delete_goods_for_requests_logistic.html'
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП', 'Менеджер']
    login_url = reverse_lazy('main:login')

    def get(self, request, *args, **kwargs):
        good = self.get_object()
        good.delete()
        goods = RequestsForLogisticsGoods.objects.filter(request=self.kwargs['request_id'])
        my_request = RequestsForLogisticsCalculations.objects.get(pk=self.kwargs['request_id'])
        return render(request, self.template_name, {'goods': goods, 'my_request': my_request})


class BidForGoodsView(MyLoginMixin, DataMixin, TemplateView):
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП', 'Менеджер']
    template_name_other = 'analytics/logistic_requests/modal/add_bid_manager.html'
    template_name_logist = 'analytics/logistic_requests/modal/add_bid.html'
    login_url = reverse_lazy('main:login')

    def get_context_data(self, *, object_list=None, **kwargs):
        if self.request.GET.get('rate'):
            rate = RequestsForLogisticsRate.objects.filter(
                good__pk=kwargs["goods_id"],
                road__pk=kwargs["road_id"],
                carrier__pk=kwargs["carrier_id"]
            )
            if rate:
                bid = rate[0]
                bid.bid = self.request.GET.get('rate')
                bid.save()
            else:
                good = RequestsForLogisticsGoods.objects.get(pk=kwargs["goods_id"])
                good.rate.create(
                    bid=self.request.GET.get('rate', ''),
                    road=RoadsList.objects.get(pk=kwargs["road_id"]),
                    carrier=CarriersList.objects.get(pk=kwargs["carrier_id"])
                )
            return dict()
        context = super().get_context_data(**kwargs)
        # Получаем товар с предварительной загрузкой связанных ставок
        good = RequestsForLogisticsGoods.objects.prefetch_related('rate').get(pk=kwargs["goods_id"])
        my_bids = good.rate.all()

        # Определяем роль пользователя
        user_role = self.request.user.role

        # Фильтрация перевозчиков и дорог в зависимости от роли
        if user_role in ['Супер Администратор', 'Логист']:
            # Для Логиста и Супер Администратора показываем все дороги и перевозчики
            roads = RoadsList.objects.all()  # Все дороги, отсортированные по ordering
            carriers = CarriersList.objects.all()  # Все перевозчики
        else:
            # Для других ролей фильтруем перевозчиков по наличию ставок
            filtered_carriers = set(bid.carrier for bid in my_bids if bid.carrier is not None)
            roads = RoadsList.objects.all()  # Все дороги, отсортированные по ordering
            carriers = list(filtered_carriers)  # Только перевозчики с ставками

        context["good"] = good
        context["my_request"] = good.request
        context["my_roads"] = good.request.roads.all()
        context['roads'] = roads
        context['carriers'] = carriers
        context['my_bids'] = my_bids
        c_def = self.get_user_context(title="Добавление ставки")
        return dict(list(context.items()) + list(c_def.items()))

    def get_template_names(self):
        if self.request.user.role == 'Логист' or self.request.user.role == 'Супер Администратор':
            return self.template_name_logist
        else:
            return self.template_name_other


def activitedBid(request, goods_id, road_id, carrier_id):
    good = RequestsForLogisticsGoods.objects.get(pk=goods_id)
    try:
        bid = RequestsForLogisticsRate.objects.filter(good__pk=goods_id, road__pk=road_id, carrier__pk=carrier_id)[0]
        good.bid = bid.bid
        good.save()
        bid.active = request.POST.get('active', True)
        bid.save()
        bids = good.rate.exclude(pk=bid.pk)
        for b in bids:
            b.active = False
            b.save()
        status = False
        for g in good.request.goods.all():
            if not g.bid:
                status = True
                break
        if status:
            good.request.status = 'Частично обработано'
        else:
            good.request.status = 'Обработано'
        good.request.save()
        template_name = 'analytics/logistic_requests/partial/edit_td_bid.html'
        return render(request, template_name, {'bid': bid})
    except:
        template_name = 'analytics/logistic_requests/partial/edit_td_bid.html'
        good.bid = ''
        if good.rate.all():
            for r in good.rate.all():
                if r.active:
                    r.active = False
                    r.save()
        good.save()
        good.request.status = 'Частично обработано'
        good.request.save()
        return render(request, template_name, {'bid': {'bid': 'Ожидание'}})


def editBidLogisticRequestsView(request, bid_id):
    bid = RequestsForLogisticsRate.objects.get(pk=bid_id)
    my_request = bid.request
    if request.POST:
        if 'rate' in request.POST:
            bid.bid = request.POST['rate']
        bid.save()
        my_request.notification = True
        my_request.status = 'Обработано'
        rate_count = my_request.rate.count()
        count = 0
        for rate in my_request.rate.all():
            if not rate.bid:
                count += 1
        if count != rate_count and count != 0:
            my_request.status = 'Частично обработано'
        elif count == rate_count:
            my_request.status = 'На просчете'
        my_request.save()
    template_name = 'analytics/logistic_requests/modal/add_bid.html'
    return render(request, template_name, {'my_request': my_request})


class DeleteFileInRequest(MyLoginMixin, DataMixin, DeleteView):
    model = RequestsForLogisticFiles
    pk_url_kwarg = 'file_id'
    context_object_name = 'my_file'
    template_name = 'analytics/logistic_requests/modal/delete_file_in_request.html'
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП', 'Менеджер']
    success_url = reverse_lazy('analytics:edit_logistic_requests')
    login_url = reverse_lazy('main:login')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Удаление файла к запросу")
        return dict(list(context.items()) + list(c_def.items()))

    def delete(self, request, *args, **kwargs):
        file = self.get_object()
        file.delete()
        return redirect('analytics:edit_logistic_requests', file.request.pk)


class DeleteLogisticRequestsView(MyLoginMixin, DataMixin, DeleteView):
    model = RequestsForLogisticsCalculations
    pk_url_kwarg = 'request_id'
    context_object_name = 'my_request'
    template_name = 'analytics/logistic_requests/modal/delete_request.html'
    role_have_perm = ['Супер Администратор', 'РОП', 'Менеджер']
    success_url = reverse_lazy('analytics:logistic_requests')
    login_url = reverse_lazy('main:login')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Удаление запроса")
        return dict(list(context.items()) + list(c_def.items()))

    def post(self, request, *args, **kwargs):
        if self.get_object().initiator.pk == self.request.user.pk:
            return super().post(request, *args, **kwargs)
        return redirect('analytics:logistic_requests')


class LogisticCarriersList(MyLoginMixin, DataMixin, ListView):
    model = CarriersList
    context_object_name = 'carriers'
    template_name = 'analytics/logistic_carriers_list/carriers_list.html'
    login_url = reverse_lazy('main:login')
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context["roads"] = RoadsList.objects.all()
        c_def = self.get_user_context(title="Список перевозчиков")
        return dict(list(context.items()) + list(c_def.items()))


class AddLogisticCarriersList(MyLoginMixin, DataMixin, CreateView):
    model = CarriersList
    form_class = AddCarriersListForm
    template_name = 'analytics/logistic_carriers_list/create_carriers_list.html'
    login_url = reverse_lazy('main:login')
    success_url = reverse_lazy('analytics:carriers_list')
    role_have_perm = ['Супер Администратор', 'Логист']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Добавление перевозчика")
        return dict(list(context.items()) + list(c_def.items()))


class EditLogisticCarriersList(MyLoginMixin, DataMixin, UpdateView):
    model = CarriersList
    form_class = EditCarriersListForm
    pk_url_kwarg = 'carrier_id'
    template_name = 'analytics/logistic_carriers_list/edit_carriers_list.html'
    login_url = reverse_lazy('main:login')
    success_url = reverse_lazy('analytics:carriers_list')
    role_have_perm = ['Супер Администратор', 'Логист']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context["carrier"] = self.get_object()
        c_def = self.get_user_context(title="Изменение данных перевозчика")
        return dict(list(context.items()) + list(c_def.items()))


class DeleteLogisticCarriersList(MyLoginMixin, DataMixin, UpdateView):
    model = CarriersList
    form_class = DeleteCarriersListForm
    template_name = 'analytics/logistic_carriers_list/delete_carriers_list.html'
    pk_url_kwarg = 'carrier_id'
    login_url = reverse_lazy('main:login')
    success_url = reverse_lazy('analytics:carriers_list')
    role_have_perm = ['Супер Администратор', 'Логист']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context["carrier"] = self.get_object()
        c_def = self.get_user_context(title="Удаление перевозчика")
        return dict(list(context.items()) + list(c_def.items()))

    def post(self, request, *args, **kwargs):
        carrier = self.get_object()
        if carrier.status:
            carrier.status = False
        carrier.save()
        return super().post(request, *args, **kwargs)


class AddRoad(MyLoginMixin, DataMixin, CreateView):
    model = RoadsList
    form_class = AddRoadForm
    template_name = 'analytics/logistic_carriers_list/create_road.html'
    login_url = reverse_lazy('main:login')
    success_url = reverse_lazy('analytics:carriers_list')
    role_have_perm = ['Супер Администратор', 'Логист']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Добавление дороги")
        return dict(list(context.items()) + list(c_def.items()))


class EditRoad(MyLoginMixin, DataMixin, UpdateView):
    model = RoadsList
    form_class = EditRoadForm
    pk_url_kwarg = 'road_id'
    template_name = 'analytics/logistic_carriers_list/edit_road.html'
    login_url = reverse_lazy('main:login')
    success_url = reverse_lazy('analytics:carriers_list')
    role_have_perm = ['Супер Администратор', 'Логист']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context["road"] = self.get_object()
        c_def = self.get_user_context(title="Изменение данных дороги")
        return dict(list(context.items()) + list(c_def.items()))


class DeleteRoad(MyLoginMixin, DataMixin, UpdateView):
    model = RoadsList
    form_class = DeleteRoadForm
    template_name = 'analytics/logistic_carriers_list/delete_road.html'
    pk_url_kwarg = 'road_id'
    login_url = reverse_lazy('main:login')
    success_url = reverse_lazy('analytics:carriers_list')
    role_have_perm = ['Супер Администратор', 'Логист']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context["road"] = self.get_object()
        c_def = self.get_user_context(title="Удаление дороги")
        return dict(list(context.items()) + list(c_def.items()))

    def post(self, request, *args, **kwargs):
        road = self.get_object()
        if road.status:
            road.status = False
        road.save()
        return super().post(request, *args, **kwargs)


def addRoadToCarriers(request, carrier_id):
    carrier = CarriersList.objects.get(pk=carrier_id)
    form = AddRoadToCarriersForm
    template_name = 'analytics/logistic_carriers_list/add_road_to_carrier.html'
    success_url = reverse_lazy('analytics:carriers_list')
    if request.POST:
        road = RoadsList.objects.get(pk=request.POST['road'])
        min_transportation_time = request.POST.get('min_transportation_time', 0)
        if len(min_transportation_time) == 0:
            min_transportation_time = 0
        max_transportation_time = request.POST.get('max_transportation_time', 0)
        if len(max_transportation_time) == 0:
            max_transportation_time = 0
        carrier.roads.add(road, through_defaults={
            "min_transportation_time": min_transportation_time,
            "max_transportation_time": max_transportation_time
        })
        return redirect(success_url)
    return render(request, template_name, {'carrier': carrier, 'form': form})


def editRoadToCarriers(request, carrier_id, road_id):
    carrier = CarriersList.objects.get(pk=carrier_id)
    road = RoadsList.objects.get(pk=road_id)
    roads_parameters = CarriersRoadParameters.objects.get(carrier=carrier, road=road)
    form = EditRoadToCarriersForm(instance=roads_parameters)
    template_name = 'analytics/logistic_carriers_list/edit_road_to_carrier.html'
    success_url = reverse_lazy('analytics:carriers_list')
    if request.POST:
        form = EditRoadToCarriersForm(request.POST)
        if form.is_valid():
            min_transportation_time = request.POST.get('min_transportation_time', 0)
            if len(min_transportation_time) == 0:
                min_transportation_time = 0
            max_transportation_time = request.POST.get('max_transportation_time', 0)
            if len(max_transportation_time) == 0:
                max_transportation_time = 0
            roads_parameters.min_transportation_time = min_transportation_time
            roads_parameters.max_transportation_time = max_transportation_time
            roads_parameters.save()
            return redirect(success_url)
    return render(request, template_name, {'carrier': carrier,
                                           'road': road,
                                           'form': form,
                                           'roads_parameters': roads_parameters})


def deleteRoadToCarriers(request, carrier_id, road_id):
    carrier = CarriersList.objects.get(pk=carrier_id)
    road = RoadsList.objects.get(pk=road_id)
    roads_parameters = CarriersRoadParameters.objects.get(carrier=carrier, road=road)
    form = DeleteRoadToCarriersForm
    template_name = 'analytics/logistic_carriers_list/delete_road_to_carrier.html'
    success_url = reverse_lazy('analytics:carriers_list')
    if request.POST:
        carrier.roads.remove(road)
        return redirect(success_url)
    return render(request, template_name, {'carrier': carrier,
                                           'road': road,
                                           'form': form,
                                           'roads_parameters': roads_parameters})


def priceListToCarrierRoad(request, carrier_id, road_id):
    carrier = CarriersList.objects.get(pk=carrier_id)
    road = RoadsList.objects.get(pk=road_id)
    roads_parameters = CarriersRoadParameters.objects.get(carrier=carrier, road=road)
    form = DeleteRoadToCarriersForm
    template_name = 'analytics/logistic_carriers_list/price_list.html'
    return render(request, template_name, {'carrier': carrier,
                                           'road': road,
                                           'form': form,
                                           'roads_parameters': roads_parameters})


def addPriceListToCarrierRoad(request, carrier_id, road_id):
    template_name = 'analytics/logistic_carriers_list/add_price_list.html'
    carrier = CarriersList.objects.get(pk=carrier_id)
    road = RoadsList.objects.get(pk=road_id)
    roads_parameters = CarriersRoadParameters.objects.get(carrier=carrier, road=road)
    density = roads_parameters.density.create()
    count = roads_parameters.density.count()
    return render(request, template_name, {'carrier': carrier,
                                           'road': road,
                                           'count': count,
                                           'density': density,
                                           'roads_parameters': roads_parameters})


def editPriceListToCarrierRoad(request, density_id):
    density = PriceListsOfCarriers.objects.get(pk=density_id)
    if request.POST:
        if request.POST.get('min_density'):
            density.min_density = request.POST.get('min_density')
        elif request.POST.get('max_density'):
            density.max_density = request.POST.get('max_density')
        elif request.POST.get('price'):
            density.price = request.POST.get('price')
        density.save()
    return render(request, 'analytics/logistic_carriers_list/edit_price_list.html',
                  {'density': density})


class LogisticCalculatorView(MyLoginMixin, DataMixin, TemplateView):
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП', 'Менеджер']
    template_name = 'analytics/calculator.html'
    login_url = reverse_lazy('main:login')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Калькулятор логистики")
        return dict(list(context.items()) + list(c_def.items()))


class CalculatorVolumeView(MyLoginMixin, DataMixin, TemplateView):
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП', 'Менеджер']
    template_name = 'analytics/calculator_volume.html'
    login_url = reverse_lazy('main:login')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Калькулятор объема")
        return dict(list(context.items()) + list(c_def.items()))


class RequestsInvoiceView(MyLoginMixin, DataMixin, TemplateView):
    role_have_perm = ['Супер Администратор', 'Логист', 'РОП', 'Менеджер', 'Бухгалтер']
    template_name = 'analytics/requests_invoice/requests_invoice.html'
    login_url = reverse_lazy('main:login')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Запрос счетов")
        return dict(list(context.items()) + list(c_def.items()))


class AddCargoView(MyLoginMixin, DataMixin, CreateView):
    model = CargoArticle
    form_class = AddCargo
    template_name = 'analytics/logistic_main/logistic_main.html'
    login_url = reverse_lazy('main:login')
    role_have_perm = ['Супер Администратор', 'Логист']
    success_url = reverse_lazy('analytics:carrier')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Ручное добавление груза")
        return dict(list(context.items()) + list(c_def.items()))

    # def form_valid(self, form):
    #     print(self.request.POST)
    #     return redirect('analytics:carrier')

    # def post(self, request, *args, **kwargs):
    #     if request.POST:
    #         print(request.POST)
    #     return redirect('analytics:carrier')


class LogisticMainView2(MyLoginMixin, DataMixin, CreateView):
    model = CargoFiles
    form_class = AddCarrierFilesForm
    template_name = 'analytics/logistic_main/logistic_main.html'
    context_object_name = 'all_articles'
    login_url = reverse_lazy('main:login')
    role_have_perm = ['Супер Администратор', 'РОП', 'Логист', 'Менеджер']
    success_url = reverse_lazy('analytics:carrier')
    message = dict()
    message['update'] = False
    factor_kg_01 = 0
    factor_kg_02 = 0
    factor_volume_10 = 0
    factor_volume_20 = 0

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_add_cargo'] = AddCargo
        context['user_pk_string'] = str(self.request.user.pk)
        context['carriers'] = CarriersList.objects.all()

        # Основной queryset
        queryset = CargoArticle.objects.select_related('responsible_manager').for_user(self.request.user)



        if self.request.user.role == 'Логист' or self.request.user.role == 'Супер Администратор':
            context['count_empty_responsible_manager'] = context['all_articles'].filter(responsible_manager=None).count()
            all_articles_count = context['all_articles'].count()
            if all_articles_count:
                if 100 - int(context['count_empty_responsible_manager'] / all_articles_count * 100) == 100 and context['count_empty_responsible_manager'] != 0:
                    context['pb_count_empty_responsible_manager'] = 99
                else:
                    context['pb_count_empty_responsible_manager'] = 100 - int(context['count_empty_responsible_manager'] / all_articles_count * 100)
            else:
                context['pb_count_empty_responsible_manager'] = 100
            context['count_empty_path_format'] = context['all_articles'].filter(path_format=None).count()
            if all_articles_count:
                if 100 - int(context['count_empty_path_format'] / all_articles_count * 100) == 100 and context['count_empty_path_format'] != 0:
                    context['pb_count_empty_path_format'] = 99
                else:
                    context['pb_count_empty_path_format'] = 100 - int(context['count_empty_path_format'] / all_articles_count * 100)
            else:
                context['pb_count_empty_path_format'] = 100
            context['all_article_with_empty_responsible_manager'] = context['all_articles'].filter(responsible_manager=None).values('article')
            context['all_article_with_empty_path_format'] = context['all_articles'].filter(path_format=None).values('article')
            context['all_articles_without_insurance'] = context['all_articles'].filter(insurance_cost__in=[None, '']).filter(time_create__year=datetime.datetime.now().year, time_create__month=datetime.datetime.now().month,
                                                                                                                             time_create__day=datetime.datetime.now().day).values('article', 'weight', 'time_from_china')
            context['new_all_articles_without_insurance'] = []
            for art in context['all_articles_without_insurance']:
                if float(art['weight'].replace(" ", "").replace(",", ".")) > 10:
                    context['new_all_articles_without_insurance'].append(art)
            context['all_articles_without_insurance'] = context['new_all_articles_without_insurance']
            context['count_articles_without_insurance'] = len(context['all_articles_without_insurance'])

        # context['count_notifications'] = context['all_articles'].filter(status='Прибыл в РФ').filter(time_cargo_release=None).count()
        context['all_articles_not_issued'] = context['all_articles'].filter(paid_by_the_client_status='Оплачено полностью').filter(time_cargo_release=None)
        context['count_all_articles_not_issued'] = context['all_articles'].filter(paid_by_the_client_status='Оплачено полностью').filter(time_cargo_release=None).count()

        if self.request.user.role == 'Менеджер':
            context['all_articles'] = context['all_articles'].filter(responsible_manager=f'{self.request.user.pk}')

        if self.request.user.role == 'РОП':
            managers_pk = [self.request.user.pk]
            context['managers'] = [{
                "pk": f"{self.request.user.pk}",
                "fi": f"{self.request.user.last_name} {self.request.user.first_name}"
            }]
            for user in CustomUser.objects.filter(role='Менеджер').filter(town=self.request.user.town).values('pk', 'last_name', 'first_name'):
                context['managers'].append({
                    "pk": f"{user['pk']}",
                    "fi": f"{user['last_name']} {user['first_name']}"
                })
                managers_pk.append(user['pk'])
            context['all_articles'] = context['all_articles'].filter(responsible_manager__in=managers_pk)
        elif self.request.user.role == 'Логист' or self.request.user.role == 'Супер Администратор':
            context['managers'] = []
            for user in CustomUser.objects.filter(role__in=['Менеджер', 'РОП'], status=True).values('pk', 'last_name', 'first_name').order_by('last_name'):
                context['managers'].append({
                    "pk": f"{user['pk']}",
                    "fi": f"{user['last_name']} {user['first_name']}"
                })

        context['count_notifications'] = context['all_articles'].filter(status='Прибыл в РФ').filter(time_cargo_release=None).count()

        if self.request.htmx:
            if self.request.htmx.trigger == 'article_search':
                q = self.request.GET.get('q') if self.request.GET.get('q') is not None else ''
                context['q'] = q
                context['all_articles'] = context['all_articles'].filter(article__iregex=q.replace(' ', ''))

        if self.request.GET.get('status') and self.request.GET.get('status') != 'Статус прибытия':
            context['status_now'] = self.request.GET.get('status')
            context['all_articles'] = context['all_articles'].filter(status=self.request.GET.get('status'))
        else:
            context['status_now'] = 'Статус прибытия'

        if self.request.GET.get('responsible_manager') and self.request.GET.get('responsible_manager') != 'Все менеджеры':
            context['responsible_manager_current'] = self.request.GET.get('responsible_manager')
            context['all_articles'] = context['all_articles'].filter(responsible_manager=context['responsible_manager_current'])
        else:
            context['responsible_manager_current'] = 'Все менеджеры'

        if self.request.GET.get('filter_date'):
            context["filter_date_now"] = self.request.GET.get('filter_date')
        else:
            context["filter_date_now"] = "Дата отправки с Китая"

        filter_for_date = 0

        if context["filter_date_now"] == 'Дата отправки с Китая':
            filter_for_date = 0
        elif context["filter_date_now"] == "Дата прибытия груза в РФ":
            filter_for_date = 1
        elif context["filter_date_now"] == "Дата выдачи груза":
            filter_for_date = 2

        if self.request.GET.get('date'):
            context['date_current'] = self.request.GET.get('date')
            if filter_for_date == 0:
                context['all_articles'] = context['all_articles'].filter(time_from_china__gte=make_aware(datetime.datetime.strptime(context['date_current'], '%Y-%m-%d')))
            elif filter_for_date == 1:
                context['all_articles'] = context['all_articles'].filter(time_cargo_arrival_to_RF__gte=make_aware(datetime.datetime.strptime(context['date_current'], '%Y-%m-%d')))
            elif filter_for_date == 2:
                context['all_articles'] = context['all_articles'].filter(time_cargo_release__gte=make_aware(datetime.datetime.strptime(context['date_current'], '%Y-%m-%d')))
        else:
            if not self.request.htmx:
                if self.request.user.role == 'Логист':
                    first_day = (datetime.datetime.now() - datetime.timedelta(days=28)).replace(day=1)
                    context['date_current'] = first_day.strftime('%Y-%m-%d')
                    if filter_for_date == 0:
                        context['all_articles'] = context['all_articles'].filter(time_from_china__gte=make_aware(first_day))
                    elif filter_for_date == 1:
                        context['all_articles'] = context['all_articles'].filter(time_cargo_arrival_to_RF__gte=make_aware(first_day))
                    elif filter_for_date == 2:
                        context['all_articles'] = context['all_articles'].filter(time_cargo_release__gte=make_aware(first_day))

        if self.request.GET.get('end_date'):
            context['end_date_current'] = self.request.GET.get('end_date')
            if filter_for_date == 0:
                context['all_articles'] = context['all_articles'].filter(time_from_china__lte=make_aware(datetime.datetime.strptime(context['end_date_current'], '%Y-%m-%d')))
            elif filter_for_date == 1:
                context['all_articles'] = context['all_articles'].filter(time_cargo_arrival_to_RF__lte=make_aware(datetime.datetime.strptime(context['end_date_current'], '%Y-%m-%d')))
            elif filter_for_date == 2:
                context['all_articles'] = context['all_articles'].filter(time_cargo_release__lte=make_aware(datetime.datetime.strptime(context['end_date_current'], '%Y-%m-%d')))

        if self.request.GET.get('day_without_payment') and self.request.GET.get('day_without_payment') != 'Просроченная оплата':
            context["day_without_payment_now"] = self.request.GET.get('day_without_payment')
            if context["day_without_payment_now"] == '1 день и больше':
                context['all_articles'] = context['all_articles'].exclude(time_cargo_release=None).annotate(day_without=F('time_paid_by_the_client_status') - F('time_cargo_release'))
                context['all_articles'] = context['all_articles'].exclude(time_cargo_release=None).annotate(day_without_today=make_aware(datetime.datetime.today()) - F('time_cargo_release'))
                context['all_articles'] = context['all_articles'].filter((Q(day_without=None) & Q(day_without_today__gte=datetime.timedelta(days=1))) | (Q(day_without__gte=datetime.timedelta(days=1))))
            elif context["day_without_payment_now"] == '7 дней и больше':
                context['all_articles'] = context['all_articles'].exclude(time_cargo_release=None).annotate(day_without=F('time_paid_by_the_client_status') - F('time_cargo_release'))
                context['all_articles'] = context['all_articles'].exclude(time_cargo_release=None).annotate(day_without_today=make_aware(datetime.datetime.today()) - F('time_cargo_release'))
                context['all_articles'] = context['all_articles'].filter((Q(day_without=None) & Q(day_without_today__gte=datetime.timedelta(days=7))) | (Q(day_without__gte=datetime.timedelta(days=7))))
        else:
            context["day_without_payment_now"] = "Просроченная оплата"

        if self.request.GET.get('paid_by_the_client') and self.request.GET.get('paid_by_the_client') != 'Оплата клиентом':
            context['paid_by_the_client_current'] = self.request.GET.get('paid_by_the_client')
            context['all_articles'] = context['all_articles'].filter(paid_by_the_client_status=context['paid_by_the_client_current'])
        else:
            context['paid_by_the_client_current'] = 'Оплата клиентом'

        if self.request.GET.get('paid_to_the_carrier') and self.request.GET.get('paid_to_the_carrier') != 'Оплата перевозчику':
            context['paid_to_the_carrier_current'] = self.request.GET.get('paid_to_the_carrier')
            context['all_articles'] = context['all_articles'].filter(payment_to_the_carrier_status=context['paid_to_the_carrier_current'])
        else:
            context['paid_to_the_carrier_current'] = 'Оплата перевозчику'

        if self.request.GET.get('carrier') and self.request.GET.get('carrier') != 'Все перевозчики':
            context['carrier_now'] = self.request.GET.get('carrier')
            context['all_articles'] = context['all_articles'].filter(carrier=self.request.GET.get('carrier'))
        else:
            context['carrier_now'] = 'Все перевозчики'

        context['form_article'] = []
        context['all_weight'] = 0
        context['all_volume'] = 0
        context['all_prr'] = 0
        context['all_tat'] = 0
        for article in context['all_articles']:
            if article.status:
                context['all_weight'] = context['all_weight'] + float(article.weight.replace(" ", "").replace(",", "."))
                context['all_volume'] = context['all_volume'] + float(article.volume.replace(" ", "").replace(",", "."))
                if article.prr:
                    context['all_prr'] = context['all_prr'] + float(article.prr.replace(" ", "").replace(",", "."))
                if article.tat_cost:
                    context['all_tat'] = context['all_tat'] + float(article.tat_cost.replace(" ", "").replace(",", "."))
        carrier_stats = (
            context['all_articles']
            .exclude(carrier__isnull=True)
            .exclude(carrier='')
            .annotate(
                weight_float=Cast('weight', FloatField()),
                volume_float=Cast('volume', FloatField())
            )
            .values('carrier')
            .annotate(
                total_weight=Sum('weight_float'),
                total_volume=Sum('volume_float'),
                total_articles=Count('id')  # ← добавляем
            )
            .order_by('carrier')  # на всякий случай
        )
        context['carrier_chart_data'] = list(carrier_stats)
        context['sum_all_weight'] = round(sum(item['total_weight'] or 0 for item in carrier_stats), 2)
        context['sum_all_volume'] = round(sum(item['total_volume'] or 0 for item in carrier_stats), 2)

        carrier_month_data = context['all_articles'] \
            .exclude(carrier__isnull=True) \
            .exclude(weight__isnull=True) \
            .annotate(month=TruncMonth('time_from_china')) \
            .values('month', 'carrier') \
            .annotate(
            total_weight=Sum(Cast('weight', FloatField())),
            total_volume=Sum(Cast('volume', FloatField())),
            count=Count('id')  # ← добавляем
        ) \
            .order_by('month')
        # Сборка в формат: { 'Янв': { 'Ян': 1234, 'Байкал': 567 }, ... }
        from collections import OrderedDict

        monthly_data = OrderedDict()
        MONTHS_RU = {
            1: 'Янв', 2: 'Фев', 3: 'Мар', 4: 'Апр', 5: 'Май', 6: 'Июн',
            7: 'Июл', 8: 'Авг', 9: 'Сен', 10: 'Окт', 11: 'Ноя', 12: 'Дек'
        }
        for entry in carrier_month_data:
            if not entry['month']: continue
            month_label = f"{MONTHS_RU[entry['month'].month]} {entry['month'].year}"  # напр. 'Янв 2024'
            carrier = entry['carrier']
            if month_label not in monthly_data:
                monthly_data[month_label] = {}
            monthly_data[month_label][carrier] = {
                'weight': round(entry['total_weight'] or 0, 2),
                'volume': round(entry['total_volume'] or 0, 2),
                'count': entry['count']  # ← включаем в данные
            }

        # Преобразуем в список для передачи в шаблон
        context['carrier_stacked_data'] = mark_safe(json.dumps(monthly_data, ensure_ascii=False))

        context['table_paginator'] = Paginator(context['all_articles'], 50)
        page_number = self.request.GET.get('page', 1)
        context['table_paginator_obj'] = context['table_paginator'].get_page(page_number)
        context['all_articles'] = context['table_paginator_obj']
        context['set_query'] = ''
        for req in self.request.GET:
            if req != 'page':
                context['set_query'] += f'&{req}={self.request.GET.get(req)}'
        context['vputi'] = 'В пути'
        context['pribil'] = 'Прибыл в РФ'
        if self.message['update']:
            context['message'] = self.message
        else:
            context['message'] = []
        self.message['update'] = False
        c_def = self.get_user_context(title="Учет грузов")
        return dict(list(context.items()) + list(c_def.items()))

    def get_template_names(self):
        if self.request.htmx:
            return "analytics/logistic_main/table.html"
        else:
            return 'analytics/logistic_main/logistic_main.html'

    def form_valid(self, form, **kwargs):
        files = self.request.FILES.getlist('file_path')  # Получаем список файлов
        self.message['success_articles'] = []
        self.message['warning_articles'] = []
        self.message['info_articles'] = []
        self.message['error'] = []

        def process_file(file_carrier):
            try:
                if file_carrier.name_carrier == 'Ян':
                    try:
                        carrier = 'Ян'
                        if settings.DEBUG:
                            dataframe = openpyxl.load_workbook(
                                os.path.join(str(os.getcwd()), 'media', str(file_carrier.file_path)), data_only=True)
                        else:
                            dataframe = openpyxl.load_workbook(
                                os.path.join(str(os.getcwd()), 'leader_cargo/media', str(file_carrier.file_path)),
                                data_only=True)
                        sheet = dataframe.get_sheet_by_name('运单')
                        address_transportation_cost = ""
                        for row in range(6, sheet.max_row):
                            if sheet[row][0].value == '送货费':
                                address_transportation_cost = float(sheet[row][13].value) / (row - 6)
                        for row in range(6, sheet.max_row):
                            if sheet[row][0].value and sheet[row][0].value != '送货费':
                                article = str(sheet[row][0].value).replace(' ', '')
                                name_goods = str(sheet[row][1].value) if sheet[row][1].value else ""
                                number_of_seats = str(sheet[row][2].value)
                                weight = str(sheet[row][4].value) if sheet[row][4].value else ""
                                volume = str(sheet[row][5].value) if sheet[row][5].value else ""
                                transportation_tariff = str(sheet[row][6].value) if sheet[row][6].value else ""
                                cost_goods = str(sheet[row][7].value) if sheet[row][7].value else ""
                                insurance_cost = str(sheet[row][8].value) if sheet[row][8].value else ""
                                packaging_cost = str(sheet[row][9].value) if sheet[row][9].value else ""
                                if sheet[row][11].value:
                                    time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][11].value, 0) if sheet[row][
                                        11].value else ""
                                else:
                                    try:
                                        time_from_china = xlrd.xldate.xldate_as_datetime(sheet[3][6].value, 0) if sheet[3][
                                            6].value else ""
                                    except TypeError:
                                        time_from_china = sheet[3][6].value
                                total_cost = str(sheet[row][13].value) if sheet[row][13].value else ""
                                try:
                                    if time_from_china >= datetime.datetime.strptime('21.11.2023', '%d.%m.%Y'):
                                        if sheet[row][6].value > 60:
                                            transportation_tariff_with_factor = float(
                                                sheet[row][6].value + self.factor_volume_10)
                                        else:
                                            transportation_tariff_with_factor = float(
                                                sheet[row][6].value + self.factor_kg_01)
                                    else:
                                        transportation_tariff_with_factor = None
                                except TypeError:
                                    transportation_tariff_with_factor = None
                                if transportation_tariff_with_factor:
                                    if sheet[row][6].value > 60:
                                        total_cost_with_factor = float(total_cost) + self.factor_volume_10 * float(volume)
                                    else:
                                        total_cost_with_factor = float(total_cost) + self.factor_kg_01 * float(weight)
                                else:
                                    total_cost_with_factor = None
                                check = False
                                old_articles = CargoArticle.objects.filter(article=article)
                                for old in old_articles:
                                    if old.article == article and old.name_goods == name_goods and old.weight == weight and old.time_from_china == make_aware(
                                            time_from_china):
                                        check = True
                                        old.carrier = carrier
                                        old.number_of_seats = number_of_seats
                                        old.volume = volume
                                        old.transportation_tariff = transportation_tariff
                                        if old.transportation_tariff_with_factor:
                                            old.transportation_tariff_with_factor = transportation_tariff_with_factor
                                            old.total_cost_with_factor = total_cost_with_factor
                                        old.cost_goods = cost_goods
                                        old.insurance_cost = insurance_cost
                                        old.packaging_cost = packaging_cost
                                        old.total_cost = total_cost
                                        old.cargo_id = file_carrier
                                        old.address_transportation_cost = address_transportation_cost
                                        old.save()
                                        self.message['info_articles'].append(
                                            f"Артикул '{old.article}' со статусом '{old.status}' "
                                            f"и датой '{(old.time_from_china + datetime.timedelta(hours=3)).strftime('%d-%m-%Y')}'")
                                        break
                                if not check:
                                    CargoArticle.objects.create(
                                        article=article,
                                        carrier=carrier,
                                        name_goods=name_goods,
                                        number_of_seats=number_of_seats,
                                        weight=weight,
                                        volume=volume,
                                        transportation_tariff=transportation_tariff,
                                        transportation_tariff_with_factor=transportation_tariff_with_factor,
                                        cost_goods=cost_goods,
                                        insurance_cost=insurance_cost,
                                        packaging_cost=packaging_cost,
                                        time_from_china=make_aware(time_from_china),
                                        total_cost=total_cost,
                                        total_cost_with_factor=total_cost_with_factor,
                                        cargo_id=file_carrier,
                                        address_transportation_cost=address_transportation_cost,
                                    )
                                    self.message['success_articles'].append(article)
                            else:
                                break
                    except:
                        carrier = 'Ян'
                        if settings.DEBUG:
                            dataframe = xlrd.open_workbook(f"{os.getcwd()}/media/{file_carrier.file_path}")
                        else:
                            dataframe = xlrd.open_workbook(f"{os.getcwd()}/leader_cargo/media/{file_carrier.file_path}")
                        sheet = dataframe.sheet_by_index(0)
                        address_transportation_cost = ""
                        data_column_text = '发货日期'
                        total_cost_text = '总计金额'
                        plus_row = 1
                        if data_column_text in sheet[2][1].value:
                            plus_row = 0
                        minus_column = 0
                        minus_column_new = 0
                        volume_column = 0
                        if total_cost_text in sheet[2 + plus_row][22].value:
                            minus_column = 1
                        if total_cost_text in sheet[2 + plus_row][21].value:
                            minus_column_new = 2
                            volume_column = 1
                        date_in_row = 1 - minus_column
                        article_number_in_row = 5 - minus_column
                        name_goods_number_in_row = 6 - minus_column
                        number_of_seats_number_in_row = 7 - minus_column
                        weight_number_in_row = 8 - minus_column
                        volume_number_in_row = 10 - minus_column - volume_column
                        transportation_tariff_number_in_row = 13 - minus_column - minus_column_new
                        transportation_tariff_volume_number_in_row = 14 - minus_column - minus_column_new
                        cost_goods_number_in_row = 15 - minus_column - minus_column_new
                        insurance_cost_number_in_row = 16 - minus_column - minus_column_new
                        packaging_cost_number_in_row = 17 - minus_column - minus_column_new
                        total_cost_number_in_row = 23 - minus_column - minus_column_new
                        for row in range(3 + plus_row, sheet.nrows):
                            if sheet[row][1].value and '合计' not in str(sheet[row][0].value):
                                article = str(sheet[row][article_number_in_row].value).replace(' ', '')
                                name_goods = str(sheet[row][name_goods_number_in_row].value) if sheet[row][
                                    name_goods_number_in_row].value else ""
                                number_of_seats = str(sheet[row][number_of_seats_number_in_row].value)
                                weight = str(sheet[row][weight_number_in_row].value) if sheet[row][
                                    weight_number_in_row].value else ""
                                volume = str(sheet[row][volume_number_in_row].value) if sheet[row][
                                    volume_number_in_row].value else ""
                                transportation_tariff = float(sheet[row][transportation_tariff_number_in_row].value) if \
                                    sheet[row][transportation_tariff_number_in_row].value else ""
                                transportation_tariff_volume = float(sheet[row][transportation_tariff_volume_number_in_row].value) if \
                                    sheet[row][transportation_tariff_volume_number_in_row].value else ""
                                cost_goods = str(sheet[row][cost_goods_number_in_row].value) if sheet[row][
                                    cost_goods_number_in_row].value else ""
                                insurance_cost = str(sheet[row][insurance_cost_number_in_row].value) if sheet[row][
                                    insurance_cost_number_in_row].value else ""
                                packaging_cost = str(sheet[row][packaging_cost_number_in_row].value) if sheet[row][
                                    packaging_cost_number_in_row].value else ""
                                if sheet[row][date_in_row].value:
                                    try:
                                        time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][date_in_row].value,
                                                                                         0) if sheet[row][date_in_row].value else ""
                                    except TypeError:
                                        time_from_china = datetime.datetime.strptime(sheet[row][date_in_row].value, '%Y-%m-%d')
                                else:
                                    try:
                                        time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][date_in_row].value,
                                                                                         0) if sheet[row][date_in_row].value else ""
                                    except TypeError:
                                        time_from_china = sheet[row][date_in_row].value
                                total_cost = str(sheet[row][total_cost_number_in_row].value) if sheet[row][
                                    total_cost_number_in_row].value else ""
                                if transportation_tariff:
                                    if transportation_tariff == 0:
                                        transportation_tariff_with_factor = float(transportation_tariff_volume) + self.factor_volume_10
                                    else:
                                        transportation_tariff_with_factor = float(transportation_tariff) + self.factor_kg_01
                                else:
                                    transportation_tariff_with_factor = float(transportation_tariff_volume) + self.factor_volume_10

                                if transportation_tariff:
                                    if transportation_tariff == 0:
                                        total_cost_with_factor = float(total_cost) + self.factor_volume_10 * float(volume)
                                    else:
                                        total_cost_with_factor = float(total_cost) + self.factor_kg_01 * float(weight)
                                else:
                                    total_cost_with_factor = float(total_cost) + self.factor_volume_10 * float(volume)

                                transportation_tariff_with_factor_multi = ''

                                if transportation_tariff and transportation_tariff_volume:
                                    transportation_tariff_with_factor_multi = f'{float(transportation_tariff) + self.factor_kg_01} + {float(transportation_tariff_volume) + self.factor_volume_20}'
                                    total_cost_with_factor = float(total_cost) + self.factor_volume_20 * float(volume) + self.factor_kg_01 * float(weight)

                                check = False
                                old_articles = CargoArticle.objects.filter(article=article)
                                for old in old_articles:
                                    if old.article == article and old.name_goods == name_goods and old.weight == weight and old.time_from_china == make_aware(
                                            time_from_china):
                                        check = True
                                        old.carrier = carrier
                                        old.number_of_seats = number_of_seats
                                        old.volume = volume
                                        old.transportation_tariff = transportation_tariff
                                        old.transportation_tariff_with_factor = transportation_tariff_with_factor
                                        old.transportation_tariff_with_factor_multi = transportation_tariff_with_factor_multi
                                        old.total_cost_with_factor = total_cost_with_factor
                                        old.cost_goods = cost_goods
                                        old.insurance_cost = insurance_cost
                                        old.packaging_cost = packaging_cost
                                        old.total_cost = total_cost
                                        old.cargo_id = file_carrier
                                        old.address_transportation_cost = address_transportation_cost
                                        old.save()
                                        self.message['info_articles'].append(
                                            f"Артикул '{old.article}' со статусом '{old.status}' "
                                            f"и датой '{(old.time_from_china + datetime.timedelta(hours=3)).strftime('%d-%m-%Y')}'")
                                        break
                                if not check:
                                    CargoArticle.objects.create(
                                        article=article,
                                        carrier=carrier,
                                        name_goods=name_goods,
                                        number_of_seats=number_of_seats,
                                        weight=weight,
                                        volume=volume,
                                        transportation_tariff=transportation_tariff,
                                        transportation_tariff_with_factor=transportation_tariff_with_factor,
                                        transportation_tariff_with_factor_multi=transportation_tariff_with_factor_multi,
                                        cost_goods=cost_goods,
                                        insurance_cost=insurance_cost,
                                        packaging_cost=packaging_cost,
                                        time_from_china=make_aware(time_from_china),
                                        total_cost=total_cost,
                                        total_cost_with_factor=total_cost_with_factor,
                                        cargo_id=file_carrier,
                                        address_transportation_cost=address_transportation_cost,
                                    )
                                    self.message['success_articles'].append(article)
                            else:
                                break
                elif file_carrier.name_carrier == 'Ян (полная машина)':
                    carrier = 'Ян'
                    if settings.DEBUG:
                        dataframe = openpyxl.load_workbook(
                            os.path.join(str(os.getcwd()), 'media', str(file_carrier.file_path)), data_only=True)
                    else:
                        dataframe = openpyxl.load_workbook(
                            os.path.join(str(os.getcwd()), 'leader_cargo/media', str(file_carrier.file_path)),
                            data_only=True)
                    sheet = dataframe.active
                    for row in range(3, sheet.max_row):
                        article = str(sheet[row][3].value).replace(' ', '')
                        number_of_seats = str(sheet[row][4].value)
                        weight = str(sheet[row][5].value) if sheet[row][5].value else ""
                        volume = str(sheet[row][6].value) if sheet[row][6].value else ""
                        address_transportation_cost = float(sheet[row][9].value)
                        try:
                            time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][0].value, 0) if sheet[row][0].value else ""
                        except TypeError:
                            time_from_china = sheet[row][0].value
                        try:
                            make_aware(time_from_china)
                        except:
                            time_from_china = datetime.datetime.strptime(time_from_china, '%Y.%m.%d')
                        total_cost = str(sheet[row][10].value) if sheet[row][10].value else ""
                        check = False
                        old_articles = CargoArticle.objects.filter(article=article)
                        for old in old_articles:
                            if old.article == article and old.status == 'В пути':
                                check = True
                                self.message['warning_articles'].append(
                                    f"Артикул '{old.article}' со статусом '{old.status}' и датой '{(old.time_from_china + datetime.timedelta(hours=3)).strftime('%d-%m-%Y')}' - уже существует")
                                break
                        if not check:
                            CargoArticle.objects.create(
                                article=article,
                                carrier=carrier,
                                number_of_seats=number_of_seats,
                                weight=weight,
                                volume=volume,
                                time_from_china=make_aware(time_from_china),
                                total_cost=total_cost,
                                cargo_id=file_carrier,
                                address_transportation_cost=address_transportation_cost,
                            )
                            self.message['success_articles'].append(article)
            except Exception as e:
                self.message['error'].append(e)

        with ThreadPoolExecutor() as executor:
            for file in files:
                file_carrier = form.save(commit=False)
                file_carrier.file_path = file
                file_carrier.save()
                executor.submit(process_file, file_carrier)
        self.message['update'] = True
        return redirect('{}?{}'.format(reverse('analytics:carrier'), self.request.META['QUERY_STRING']))


class LogisticMainView(MyLoginMixin, DataMixin, CreateView):
    model = CargoFiles
    form_class = AddCarrierFilesForm
    template_name = 'analytics/logistic_main/logistic_main.html'
    context_object_name = 'all_articles'
    login_url = reverse_lazy('main:login')
    role_have_perm = ['Супер Администратор', 'РОП', 'Логист', 'Менеджер']
    success_url = reverse_lazy('analytics:carrier')
    message = dict()
    message['update'] = False
    factor_kg_01 = 0
    factor_kg_02 = 0
    factor_volume_10 = 0
    factor_volume_20 = 0

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_add_cargo'] = AddCargo
        context['user_pk_string'] = str(self.request.user.pk)
        context['all_articles'] = CargoArticle.objects.all().select_related('responsible_manager')
        context['carriers'] = CarriersList.objects.all()
        if self.request.user.role == 'Логист' or self.request.user.role == 'Супер Администратор':
            context['count_empty_responsible_manager'] = context['all_articles'].filter(responsible_manager=None).count()
            all_articles_count = context['all_articles'].count()
            if all_articles_count:
                if 100 - int(context['count_empty_responsible_manager'] / all_articles_count * 100) == 100 and context['count_empty_responsible_manager'] != 0:
                    context['pb_count_empty_responsible_manager'] = 99
                else:
                    context['pb_count_empty_responsible_manager'] = 100 - int(context['count_empty_responsible_manager'] / all_articles_count * 100)
            else:
                context['pb_count_empty_responsible_manager'] = 100
            context['count_empty_path_format'] = context['all_articles'].filter(path_format=None).count()
            if all_articles_count:
                if 100 - int(context['count_empty_path_format'] / all_articles_count * 100) == 100 and context['count_empty_path_format'] != 0:
                    context['pb_count_empty_path_format'] = 99
                else:
                    context['pb_count_empty_path_format'] = 100 - int(context['count_empty_path_format'] / all_articles_count * 100)
            else:
                context['pb_count_empty_path_format'] = 100
            context['all_article_with_empty_responsible_manager'] = context['all_articles'].filter(responsible_manager=None).values('article')
            context['all_article_with_empty_path_format'] = context['all_articles'].filter(path_format=None).values('article')
            context['all_articles_without_insurance'] = context['all_articles'].filter(insurance_cost__in=[None, '']).filter(time_create__year=datetime.datetime.now().year, time_create__month=datetime.datetime.now().month,
                                                                                                                             time_create__day=datetime.datetime.now().day).values('article', 'weight', 'time_from_china')
            context['new_all_articles_without_insurance'] = []
            for art in context['all_articles_without_insurance']:
                if float(art['weight'].replace(" ", "").replace(",", ".")) > 10:
                    context['new_all_articles_without_insurance'].append(art)
            context['all_articles_without_insurance'] = context['new_all_articles_without_insurance']
            context['count_articles_without_insurance'] = len(context['all_articles_without_insurance'])

        # context['count_notifications'] = context['all_articles'].filter(status='Прибыл в РФ').filter(time_cargo_release=None).count()
        context['all_articles_not_issued'] = context['all_articles'].filter(paid_by_the_client_status='Оплачено полностью').filter(time_cargo_release=None)
        context['count_all_articles_not_issued'] = context['all_articles'].filter(paid_by_the_client_status='Оплачено полностью').filter(time_cargo_release=None).count()

        if self.request.user.role == 'Менеджер':
            context['all_articles'] = context['all_articles'].filter(responsible_manager=f'{self.request.user.pk}')


        if self.request.user.role == 'Логист' or self.request.user.role == 'Супер Администратор' or self.request.user.role == 'РОП':
            context['managers'] = []
            for user in CustomUser.objects.filter(role__in=['Менеджер', 'РОП'], status=True).values('pk', 'last_name', 'first_name').order_by('last_name'):
                context['managers'].append({
                    "pk": f"{user['pk']}",
                    "fi": f"{user['last_name']} {user['first_name']}"
                })

        context['count_notifications'] = context['all_articles'].filter(status='Прибыл в РФ').filter(time_cargo_release=None).count()

        if self.request.htmx:
            if self.request.htmx.trigger == 'article_search':
                q = self.request.GET.get('q') if self.request.GET.get('q') is not None else ''
                context['q'] = q
                context['all_articles'] = context['all_articles'].filter(article__iregex=q.replace(' ', ''))

        if self.request.GET.get('status') and self.request.GET.get('status') != 'Статус прибытия':
            context['status_now'] = self.request.GET.get('status')
            context['all_articles'] = context['all_articles'].filter(status=self.request.GET.get('status'))
        else:
            context['status_now'] = 'Статус прибытия'

        if self.request.GET.get('responsible_manager') and self.request.GET.get('responsible_manager') != 'Все менеджеры':
            context['responsible_manager_current'] = self.request.GET.get('responsible_manager')
            context['all_articles'] = context['all_articles'].filter(responsible_manager=context['responsible_manager_current'])
        else:
            context['responsible_manager_current'] = 'Все менеджеры'

        if self.request.GET.get('filter_date'):
            context["filter_date_now"] = self.request.GET.get('filter_date')
        else:
            context["filter_date_now"] = "Дата отправки с Китая"

        filter_for_date = 0

        if context["filter_date_now"] == 'Дата отправки с Китая':
            filter_for_date = 0
        elif context["filter_date_now"] == "Дата прибытия груза в РФ":
            filter_for_date = 1
        elif context["filter_date_now"] == "Дата выдачи груза":
            filter_for_date = 2

        if self.request.GET.get('date'):
            context['date_current'] = self.request.GET.get('date')
            if filter_for_date == 0:
                context['all_articles'] = context['all_articles'].filter(time_from_china__gte=make_aware(datetime.datetime.strptime(context['date_current'], '%Y-%m-%d')))
            elif filter_for_date == 1:
                context['all_articles'] = context['all_articles'].filter(time_cargo_arrival_to_RF__gte=make_aware(datetime.datetime.strptime(context['date_current'], '%Y-%m-%d')))
            elif filter_for_date == 2:
                context['all_articles'] = context['all_articles'].filter(time_cargo_release__gte=make_aware(datetime.datetime.strptime(context['date_current'], '%Y-%m-%d')))
        else:
            if not self.request.htmx:
                if self.request.user.role == 'Логист':
                    first_day = (datetime.datetime.now() - datetime.timedelta(days=28)).replace(day=1)
                    context['date_current'] = first_day.strftime('%Y-%m-%d')
                    if filter_for_date == 0:
                        context['all_articles'] = context['all_articles'].filter(time_from_china__gte=make_aware(first_day))
                    elif filter_for_date == 1:
                        context['all_articles'] = context['all_articles'].filter(time_cargo_arrival_to_RF__gte=make_aware(first_day))
                    elif filter_for_date == 2:
                        context['all_articles'] = context['all_articles'].filter(time_cargo_release__gte=make_aware(first_day))

        if self.request.GET.get('end_date'):
            context['end_date_current'] = self.request.GET.get('end_date')
            if filter_for_date == 0:
                context['all_articles'] = context['all_articles'].filter(time_from_china__lte=make_aware(datetime.datetime.strptime(context['end_date_current'], '%Y-%m-%d')))
            elif filter_for_date == 1:
                context['all_articles'] = context['all_articles'].filter(time_cargo_arrival_to_RF__lte=make_aware(datetime.datetime.strptime(context['end_date_current'], '%Y-%m-%d')))
            elif filter_for_date == 2:
                context['all_articles'] = context['all_articles'].filter(time_cargo_release__lte=make_aware(datetime.datetime.strptime(context['end_date_current'], '%Y-%m-%d')))

        if self.request.GET.get('day_without_payment') and self.request.GET.get('day_without_payment') != 'Просроченная оплата':
            context["day_without_payment_now"] = self.request.GET.get('day_without_payment')
            if context["day_without_payment_now"] == '1 день и больше':
                context['all_articles'] = context['all_articles'].exclude(time_cargo_release=None).annotate(day_without=F('time_paid_by_the_client_status') - F('time_cargo_release'))
                context['all_articles'] = context['all_articles'].exclude(time_cargo_release=None).annotate(day_without_today=make_aware(datetime.datetime.today()) - F('time_cargo_release'))
                context['all_articles'] = context['all_articles'].filter((Q(day_without=None) & Q(day_without_today__gte=datetime.timedelta(days=1))) | (Q(day_without__gte=datetime.timedelta(days=1))))
            elif context["day_without_payment_now"] == '7 дней и больше':
                context['all_articles'] = context['all_articles'].exclude(time_cargo_release=None).annotate(day_without=F('time_paid_by_the_client_status') - F('time_cargo_release'))
                context['all_articles'] = context['all_articles'].exclude(time_cargo_release=None).annotate(day_without_today=make_aware(datetime.datetime.today()) - F('time_cargo_release'))
                context['all_articles'] = context['all_articles'].filter((Q(day_without=None) & Q(day_without_today__gte=datetime.timedelta(days=7))) | (Q(day_without__gte=datetime.timedelta(days=7))))
        else:
            context["day_without_payment_now"] = "Просроченная оплата"

        if self.request.GET.get('paid_by_the_client') and self.request.GET.get('paid_by_the_client') != 'Оплата клиентом':
            context['paid_by_the_client_current'] = self.request.GET.get('paid_by_the_client')
            context['all_articles'] = context['all_articles'].filter(paid_by_the_client_status=context['paid_by_the_client_current'])
        else:
            context['paid_by_the_client_current'] = 'Оплата клиентом'

        if self.request.GET.get('paid_to_the_carrier') and self.request.GET.get('paid_to_the_carrier') != 'Оплата перевозчику':
            context['paid_to_the_carrier_current'] = self.request.GET.get('paid_to_the_carrier')
            context['all_articles'] = context['all_articles'].filter(payment_to_the_carrier_status=context['paid_to_the_carrier_current'])
        else:
            context['paid_to_the_carrier_current'] = 'Оплата перевозчику'

        if self.request.GET.get('carrier') and self.request.GET.get('carrier') != 'Все перевозчики':
            context['carrier_now'] = self.request.GET.get('carrier')
            context['all_articles'] = context['all_articles'].filter(carrier=self.request.GET.get('carrier'))
        else:
            context['carrier_now'] = 'Все перевозчики'

        context['form_article'] = []
        context['all_weight'] = 0
        context['all_volume'] = 0
        context['all_prr'] = 0
        context['all_tat'] = 0
        for article in context['all_articles']:
            if article.status:
                context['all_weight'] = context['all_weight'] + float(article.weight.replace(" ", "").replace(",", "."))
                context['all_volume'] = context['all_volume'] + float(article.volume.replace(" ", "").replace(",", "."))
                if article.prr:
                    context['all_prr'] = context['all_prr'] + float(article.prr.replace(" ", "").replace(",", "."))
                if article.tat_cost:
                    context['all_tat'] = context['all_tat'] + float(article.tat_cost.replace(" ", "").replace(",", "."))
        carrier_stats = (
            context['all_articles']
            .exclude(carrier__isnull=True)
            .exclude(carrier='')
            .annotate(
                weight_float=Cast('weight', FloatField()),
                volume_float=Cast('volume', FloatField())
            )
            .values('carrier')
            .annotate(
                total_weight=Sum('weight_float'),
                total_volume=Sum('volume_float'),
                total_articles=Count('id')  # ← добавляем
            )
            .order_by('carrier')  # на всякий случай
        )
        context['carrier_chart_data'] = list(carrier_stats)
        context['sum_all_weight'] = round(sum(item['total_weight'] or 0 for item in carrier_stats), 2)
        context['sum_all_volume'] = round(sum(item['total_volume'] or 0 for item in carrier_stats), 2)

        carrier_month_data = context['all_articles'] \
            .exclude(carrier__isnull=True) \
            .exclude(weight__isnull=True) \
            .annotate(month=TruncMonth('time_from_china')) \
            .values('month', 'carrier') \
            .annotate(
            total_weight=Sum(Cast('weight', FloatField())),
            total_volume=Sum(Cast('volume', FloatField())),
            count=Count('id')  # ← добавляем
        ) \
            .order_by('month')
        # Сборка в формат: { 'Янв': { 'Ян': 1234, 'Байкал': 567 }, ... }
        from collections import OrderedDict

        monthly_data = OrderedDict()
        MONTHS_RU = {
            1: 'Янв', 2: 'Фев', 3: 'Мар', 4: 'Апр', 5: 'Май', 6: 'Июн',
            7: 'Июл', 8: 'Авг', 9: 'Сен', 10: 'Окт', 11: 'Ноя', 12: 'Дек'
        }
        for entry in carrier_month_data:
            if not entry['month']: continue
            month_label = f"{MONTHS_RU[entry['month'].month]} {entry['month'].year}"  # напр. 'Янв 2024'
            carrier = entry['carrier']
            if month_label not in monthly_data:
                monthly_data[month_label] = {}
            monthly_data[month_label][carrier] = {
                'weight': round(entry['total_weight'] or 0, 2),
                'volume': round(entry['total_volume'] or 0, 2),
                'count': entry['count']  # ← включаем в данные
            }

        # Преобразуем в список для передачи в шаблон
        context['carrier_stacked_data'] = mark_safe(json.dumps(monthly_data, ensure_ascii=False))

        context['table_paginator'] = Paginator(context['all_articles'], 50)
        page_number = self.request.GET.get('page', 1)
        context['table_paginator_obj'] = context['table_paginator'].get_page(page_number)
        context['all_articles'] = context['table_paginator_obj']
        context['set_query'] = ''
        for req in self.request.GET:
            if req != 'page':
                context['set_query'] += f'&{req}={self.request.GET.get(req)}'
        context['vputi'] = 'В пути'
        context['pribil'] = 'Прибыл в РФ'
        if self.message['update']:
            context['message'] = self.message
        else:
            context['message'] = []
        self.message['update'] = False
        c_def = self.get_user_context(title="Учет грузов")
        return dict(list(context.items()) + list(c_def.items()))

    def get_template_names(self):
        if self.request.htmx:
            return "analytics/logistic_main/table.html"
        else:
            return 'analytics/logistic_main/logistic_main.html'

    def form_valid(self, form, **kwargs):
        files = self.request.FILES.getlist('file_path')  # Получаем список файлов
        self.message['success_articles'] = []
        self.message['warning_articles'] = []
        self.message['info_articles'] = []
        self.message['error'] = []

        def process_file(file_carrier):
            try:
                if file_carrier.name_carrier == 'Ян':
                    try:
                        carrier = 'Ян'
                        if settings.DEBUG:
                            dataframe = openpyxl.load_workbook(
                                os.path.join(str(os.getcwd()), 'media', str(file_carrier.file_path)), data_only=True)
                        else:
                            dataframe = openpyxl.load_workbook(
                                os.path.join(str(os.getcwd()), 'leader_cargo/media', str(file_carrier.file_path)),
                                data_only=True)
                        sheet = dataframe.get_sheet_by_name('运单')
                        address_transportation_cost = ""
                        for row in range(6, sheet.max_row):
                            if sheet[row][0].value == '送货费':
                                address_transportation_cost = float(sheet[row][13].value) / (row - 6)
                        for row in range(6, sheet.max_row):
                            if sheet[row][0].value and sheet[row][0].value != '送货费':
                                article = str(sheet[row][0].value).replace(' ', '')
                                name_goods = str(sheet[row][1].value) if sheet[row][1].value else ""
                                number_of_seats = str(sheet[row][2].value)
                                weight = str(sheet[row][4].value) if sheet[row][4].value else ""
                                volume = str(sheet[row][5].value) if sheet[row][5].value else ""
                                transportation_tariff = str(sheet[row][6].value) if sheet[row][6].value else ""
                                cost_goods = str(sheet[row][7].value) if sheet[row][7].value else ""
                                insurance_cost = str(sheet[row][8].value) if sheet[row][8].value else ""
                                packaging_cost = str(sheet[row][9].value) if sheet[row][9].value else ""
                                if sheet[row][11].value:
                                    time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][11].value, 0) if sheet[row][
                                        11].value else ""
                                else:
                                    try:
                                        time_from_china = xlrd.xldate.xldate_as_datetime(sheet[3][6].value, 0) if sheet[3][
                                            6].value else ""
                                    except TypeError:
                                        time_from_china = sheet[3][6].value
                                total_cost = str(sheet[row][13].value) if sheet[row][13].value else ""
                                try:
                                    if time_from_china >= datetime.datetime.strptime('21.11.2023', '%d.%m.%Y'):
                                        if sheet[row][6].value > 60:
                                            transportation_tariff_with_factor = float(
                                                sheet[row][6].value + self.factor_volume_10)
                                        else:
                                            transportation_tariff_with_factor = float(
                                                sheet[row][6].value + self.factor_kg_01)
                                    else:
                                        transportation_tariff_with_factor = None
                                except TypeError:
                                    transportation_tariff_with_factor = None
                                if transportation_tariff_with_factor:
                                    if sheet[row][6].value > 60:
                                        total_cost_with_factor = float(total_cost) + self.factor_volume_10 * float(volume)
                                    else:
                                        total_cost_with_factor = float(total_cost) + self.factor_kg_01 * float(weight)
                                else:
                                    total_cost_with_factor = None
                                check = False
                                old_articles = CargoArticle.objects.filter(article=article)
                                for old in old_articles:
                                    if old.article == article and old.name_goods == name_goods and old.weight == weight and old.time_from_china == make_aware(
                                            time_from_china):
                                        check = True
                                        old.carrier = carrier
                                        old.number_of_seats = number_of_seats
                                        old.volume = volume
                                        old.transportation_tariff = transportation_tariff
                                        if old.transportation_tariff_with_factor:
                                            old.transportation_tariff_with_factor = transportation_tariff_with_factor
                                            old.total_cost_with_factor = total_cost_with_factor
                                        old.cost_goods = cost_goods
                                        old.insurance_cost = insurance_cost
                                        old.packaging_cost = packaging_cost
                                        old.total_cost = total_cost
                                        old.cargo_id = file_carrier
                                        old.address_transportation_cost = address_transportation_cost
                                        old.save()
                                        self.message['info_articles'].append(
                                            f"Артикул '{old.article}' со статусом '{old.status}' "
                                            f"и датой '{(old.time_from_china + datetime.timedelta(hours=3)).strftime('%d-%m-%Y')}'")
                                        break
                                if not check:
                                    CargoArticle.objects.create(
                                        article=article,
                                        carrier=carrier,
                                        name_goods=name_goods,
                                        number_of_seats=number_of_seats,
                                        weight=weight,
                                        volume=volume,
                                        transportation_tariff=transportation_tariff,
                                        transportation_tariff_with_factor=transportation_tariff_with_factor,
                                        cost_goods=cost_goods,
                                        insurance_cost=insurance_cost,
                                        packaging_cost=packaging_cost,
                                        time_from_china=make_aware(time_from_china),
                                        total_cost=total_cost,
                                        total_cost_with_factor=total_cost_with_factor,
                                        cargo_id=file_carrier,
                                        address_transportation_cost=address_transportation_cost,
                                    )
                                    self.message['success_articles'].append(article)
                            else:
                                break
                    except:
                        carrier = 'Ян'
                        if settings.DEBUG:
                            dataframe = xlrd.open_workbook(f"{os.getcwd()}/media/{file_carrier.file_path}")
                        else:
                            dataframe = xlrd.open_workbook(f"{os.getcwd()}/leader_cargo/media/{file_carrier.file_path}")
                        sheet = dataframe.sheet_by_index(0)
                        address_transportation_cost = ""
                        data_column_text = '发货日期'
                        total_cost_text = '总计金额'
                        plus_row = 1
                        if data_column_text in sheet[2][1].value:
                            plus_row = 0
                        minus_column = 0
                        minus_column_new = 0
                        volume_column = 0
                        if total_cost_text in sheet[2 + plus_row][22].value:
                            minus_column = 1
                        if total_cost_text in sheet[2 + plus_row][21].value:
                            minus_column_new = 2
                            volume_column = 1
                        date_in_row = 1 - minus_column
                        article_number_in_row = 5 - minus_column
                        name_goods_number_in_row = 6 - minus_column
                        number_of_seats_number_in_row = 7 - minus_column
                        weight_number_in_row = 8 - minus_column
                        volume_number_in_row = 10 - minus_column - volume_column
                        transportation_tariff_number_in_row = 13 - minus_column - minus_column_new
                        transportation_tariff_volume_number_in_row = 14 - minus_column - minus_column_new
                        cost_goods_number_in_row = 15 - minus_column - minus_column_new
                        insurance_cost_number_in_row = 16 - minus_column - minus_column_new
                        packaging_cost_number_in_row = 17 - minus_column - minus_column_new
                        total_cost_number_in_row = 23 - minus_column - minus_column_new
                        for row in range(3 + plus_row, sheet.nrows):
                            if sheet[row][1].value and '合计' not in str(sheet[row][0].value):
                                article = str(sheet[row][article_number_in_row].value).replace(' ', '')
                                name_goods = str(sheet[row][name_goods_number_in_row].value) if sheet[row][
                                    name_goods_number_in_row].value else ""
                                number_of_seats = str(sheet[row][number_of_seats_number_in_row].value)
                                weight = str(sheet[row][weight_number_in_row].value) if sheet[row][
                                    weight_number_in_row].value else ""
                                volume = str(sheet[row][volume_number_in_row].value) if sheet[row][
                                    volume_number_in_row].value else ""
                                transportation_tariff = float(sheet[row][transportation_tariff_number_in_row].value) if \
                                    sheet[row][transportation_tariff_number_in_row].value else ""
                                transportation_tariff_volume = float(sheet[row][transportation_tariff_volume_number_in_row].value) if \
                                    sheet[row][transportation_tariff_volume_number_in_row].value else ""
                                cost_goods = str(sheet[row][cost_goods_number_in_row].value) if sheet[row][
                                    cost_goods_number_in_row].value else ""
                                insurance_cost = str(sheet[row][insurance_cost_number_in_row].value) if sheet[row][
                                    insurance_cost_number_in_row].value else ""
                                packaging_cost = str(sheet[row][packaging_cost_number_in_row].value) if sheet[row][
                                    packaging_cost_number_in_row].value else ""
                                if sheet[row][date_in_row].value:
                                    try:
                                        time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][date_in_row].value,
                                                                                         0) if sheet[row][date_in_row].value else ""
                                    except TypeError:
                                        time_from_china = datetime.datetime.strptime(sheet[row][date_in_row].value, '%Y-%m-%d')
                                else:
                                    try:
                                        time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][date_in_row].value,
                                                                                         0) if sheet[row][date_in_row].value else ""
                                    except TypeError:
                                        time_from_china = sheet[row][date_in_row].value
                                total_cost = str(sheet[row][total_cost_number_in_row].value) if sheet[row][
                                    total_cost_number_in_row].value else ""
                                if transportation_tariff:
                                    if transportation_tariff == 0:
                                        transportation_tariff_with_factor = float(transportation_tariff_volume) + self.factor_volume_10
                                    else:
                                        transportation_tariff_with_factor = float(transportation_tariff) + self.factor_kg_01
                                else:
                                    transportation_tariff_with_factor = float(transportation_tariff_volume) + self.factor_volume_10

                                if transportation_tariff:
                                    if transportation_tariff == 0:
                                        total_cost_with_factor = float(total_cost) + self.factor_volume_10 * float(volume)
                                    else:
                                        total_cost_with_factor = float(total_cost) + self.factor_kg_01 * float(weight)
                                else:
                                    total_cost_with_factor = float(total_cost) + self.factor_volume_10 * float(volume)

                                transportation_tariff_with_factor_multi = ''

                                if transportation_tariff and transportation_tariff_volume:
                                    transportation_tariff_with_factor_multi = f'{float(transportation_tariff) + self.factor_kg_01} + {float(transportation_tariff_volume) + self.factor_volume_20}'
                                    total_cost_with_factor = float(total_cost) + self.factor_volume_20 * float(volume) + self.factor_kg_01 * float(weight)

                                check = False
                                old_articles = CargoArticle.objects.filter(article=article)
                                for old in old_articles:
                                    if old.article == article and old.name_goods == name_goods and old.weight == weight and old.time_from_china == make_aware(
                                            time_from_china):
                                        check = True
                                        old.carrier = carrier
                                        old.number_of_seats = number_of_seats
                                        old.volume = volume
                                        old.transportation_tariff = transportation_tariff
                                        old.transportation_tariff_with_factor = transportation_tariff_with_factor
                                        old.transportation_tariff_with_factor_multi = transportation_tariff_with_factor_multi
                                        old.total_cost_with_factor = total_cost_with_factor
                                        old.cost_goods = cost_goods
                                        old.insurance_cost = insurance_cost
                                        old.packaging_cost = packaging_cost
                                        old.total_cost = total_cost
                                        old.cargo_id = file_carrier
                                        old.address_transportation_cost = address_transportation_cost
                                        old.save()
                                        self.message['info_articles'].append(
                                            f"Артикул '{old.article}' со статусом '{old.status}' "
                                            f"и датой '{(old.time_from_china + datetime.timedelta(hours=3)).strftime('%d-%m-%Y')}'")
                                        break
                                if not check:
                                    CargoArticle.objects.create(
                                        article=article,
                                        carrier=carrier,
                                        name_goods=name_goods,
                                        number_of_seats=number_of_seats,
                                        weight=weight,
                                        volume=volume,
                                        transportation_tariff=transportation_tariff,
                                        transportation_tariff_with_factor=transportation_tariff_with_factor,
                                        transportation_tariff_with_factor_multi=transportation_tariff_with_factor_multi,
                                        cost_goods=cost_goods,
                                        insurance_cost=insurance_cost,
                                        packaging_cost=packaging_cost,
                                        time_from_china=make_aware(time_from_china),
                                        total_cost=total_cost,
                                        total_cost_with_factor=total_cost_with_factor,
                                        cargo_id=file_carrier,
                                        address_transportation_cost=address_transportation_cost,
                                    )
                                    self.message['success_articles'].append(article)
                            else:
                                break
                elif file_carrier.name_carrier == 'Ян (полная машина)':
                    carrier = 'Ян'
                    if settings.DEBUG:
                        dataframe = openpyxl.load_workbook(
                            os.path.join(str(os.getcwd()), 'media', str(file_carrier.file_path)), data_only=True)
                    else:
                        dataframe = openpyxl.load_workbook(
                            os.path.join(str(os.getcwd()), 'leader_cargo/media', str(file_carrier.file_path)),
                            data_only=True)
                    sheet = dataframe.active
                    for row in range(3, sheet.max_row):
                        article = str(sheet[row][3].value).replace(' ', '')
                        number_of_seats = str(sheet[row][4].value)
                        weight = str(sheet[row][5].value) if sheet[row][5].value else ""
                        volume = str(sheet[row][6].value) if sheet[row][6].value else ""
                        address_transportation_cost = float(sheet[row][9].value)
                        try:
                            time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][0].value, 0) if sheet[row][0].value else ""
                        except TypeError:
                            time_from_china = sheet[row][0].value
                        try:
                            make_aware(time_from_china)
                        except:
                            time_from_china = datetime.datetime.strptime(time_from_china, '%Y.%m.%d')
                        total_cost = str(sheet[row][10].value) if sheet[row][10].value else ""
                        check = False
                        old_articles = CargoArticle.objects.filter(article=article)
                        for old in old_articles:
                            if old.article == article and old.status == 'В пути':
                                check = True
                                self.message['warning_articles'].append(
                                    f"Артикул '{old.article}' со статусом '{old.status}' и датой '{(old.time_from_china + datetime.timedelta(hours=3)).strftime('%d-%m-%Y')}' - уже существует")
                                break
                        if not check:
                            CargoArticle.objects.create(
                                article=article,
                                carrier=carrier,
                                number_of_seats=number_of_seats,
                                weight=weight,
                                volume=volume,
                                time_from_china=make_aware(time_from_china),
                                total_cost=total_cost,
                                cargo_id=file_carrier,
                                address_transportation_cost=address_transportation_cost,
                            )
                            self.message['success_articles'].append(article)
            except Exception as e:
                self.message['error'].append(e)

        with ThreadPoolExecutor() as executor:
            for file in files:
                file_carrier = form.save(commit=False)
                file_carrier.file_path = file
                file_carrier.save()
                executor.submit(process_file, file_carrier)
        self.message['update'] = True
        return redirect('{}?{}'.format(reverse('analytics:carrier'), self.request.META['QUERY_STRING']))
# elif file_carrier.name_carrier == 'Валька':
# carrier = 'Валька'
# if settings.DEBUG:
#     dataframe = openpyxl.load_workbook(f"{os.getcwd()}/media/{file_carrier.file_path}", data_only=True)
# else:
#     dataframe = openpyxl.load_workbook(f"{os.getcwd()}/leader_cargo/media/{file_carrier.file_path}",
#                                        data_only=True)
# sheet = dataframe.active
# for row in range(2, sheet.max_row):
#     if sheet[row][4].value and sheet[row][3].value and sheet[row][2].value:
#         article = str(sheet[row][4].value).replace(' ', '')
#         name_goods = str(sheet[row][6].value) if sheet[row][6].value else ""
#         number_of_seats = str(sheet[row][7].value)
#         weight = str(sheet[row][8].value) if sheet[row][8].value else ""
#         volume = str(sheet[row][9].value) if sheet[row][9].value else ""
#         transportation_tariff = str(round(float(sheet[row][12].value) / float(sheet[row][8].value), 2))
#         cost_goods = str(sheet[row][13].value) if sheet[row][13].value else ""
#         insurance_cost = str(sheet[row][14].value) if sheet[row][14].value else ""
#         packaging_cost = str(sheet[row][15].value) if sheet[row][15].value else ""
#         try:
#             time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][3].value, 0) if sheet[row][
#                 3].value else ""
#         except TypeError:
#             time_from_china = sheet[row][3].value
#         total_cost = str(sheet[row][16].value) if sheet[row][16].value else ""
#         try:
#             if time_from_china >= datetime.datetime.strptime('21.11.2023', '%d.%m.%Y'):
#                 if round(float(sheet[row][12].value) / float(sheet[row][8].value), 2) > 60:
#                     transportation_tariff_with_factor = float(
#                         round(float(sheet[row][12].value) / float(sheet[row][8].value),
#                               2) + self.factor_volume_20)
#                 else:
#                     transportation_tariff_with_factor = float(
#                         round(float(sheet[row][12].value) / float(sheet[row][8].value),
#                               2) + self.factor_kg_02)
#             else:
#                 transportation_tariff_with_factor = None
#         except TypeError:
#             transportation_tariff_with_factor = None
#         if transportation_tariff_with_factor:
#             if round(float(sheet[row][12].value) / float(sheet[row][8].value), 2) > 60:
#                 total_cost_with_factor = float(total_cost) + self.factor_volume_20 * float(volume)
#             else:
#                 total_cost_with_factor = float(total_cost) + self.factor_kg_02 * float(weight)
#         else:
#             total_cost_with_factor = None
#         check = False
#         old_articles = CargoArticle.objects.filter(article=article)
#         for old in old_articles:
#             if old.article == article and old.name_goods == name_goods and old.number_of_seats == number_of_seats and old.cost_goods == cost_goods and old.insurance_cost == insurance_cost \
#                     and old.weight == weight and old.volume == volume and old.transportation_tariff == transportation_tariff and old.packaging_cost == packaging_cost \
#                     and old.time_from_china == make_aware(
#                 time_from_china) and old.total_cost == total_cost:
#                 check = True
#                 break
#         if not check:
#             CargoArticle.objects.create(
#                 article=article,
#                 carrier=carrier,
#                 name_goods=name_goods,
#                 number_of_seats=number_of_seats,
#                 weight=weight,
#                 volume=volume,
#                 transportation_tariff=transportation_tariff,
#                 transportation_tariff_with_factor=transportation_tariff_with_factor,
#                 cost_goods=cost_goods,
#                 insurance_cost=insurance_cost,
#                 packaging_cost=packaging_cost,
#                 time_from_china=make_aware(time_from_china),
#                 total_cost=total_cost,
#                 total_cost_with_factor=total_cost_with_factor,
#                 cargo_id=file_carrier,
#             )
#             self.message['success_articles'].append(article)
#     else:
#         break
# elif file_carrier.name_carrier == 'Мурад':
# carrier = 'Мурад'
# if settings.DEBUG:
#     dataframe = openpyxl.load_workbook(f"{os.getcwd()}/media/{file_carrier.file_path}", data_only=True)
# else:
#     dataframe = openpyxl.load_workbook(f"{os.getcwd()}/leader_cargo/media/{file_carrier.file_path}",
#                                        data_only=True)
# sheet = dataframe.active
# for row in range(2, sheet.max_row + 1):
#     if sheet[row][5].value:
#         if sheet[row][5].value.find("（") != -1:
#             article = str(sheet[row][5].value.split("（")[0].replace(" ", "")) + '\n' + str(
#                 sheet[row][3].value)
#             if sheet[row][5].value.find("）") != -1:
#                 name_goods = str(sheet[row][5].value.split("（")[1].replace(" ", "").replace("）", ""))
#             else:
#                 name_goods = str(sheet[row][5].value.split("（")[1].replace(" ", "").replace(")", ""))
#         else:
#             article = str(sheet[row][5].value.split("(")[0].replace(" ", "")) + '\n' + str(
#                 sheet[row][3].value)
#             if sheet[row][5].value.find("）") != -1:
#                 name_goods = str(sheet[row][5].value.split("(")[1].replace(" ", "").replace("）", ""))
#             else:
#                 name_goods = str(sheet[row][5].value.split("(")[1].replace(" ", "").replace(")", ""))
#         number_of_seats = str(sheet[row][4].value)
#         weight = str(sheet[row][6].value) if sheet[row][4].value else ""
#         volume = str(sheet[row][7].value) if sheet[row][5].value else ""
#         transportation_tariff = str(sheet[row][15].value) if sheet[row][15].value else ""
#         cost_goods = str(sheet[row][12].value) if sheet[row][12].value else ""
#         insurance_cost = str(sheet[row][13].value) if sheet[row][13].value else ""
#         packaging_cost = str(sheet[row][14].value) if sheet[row][14].value else ""
#         try:
#             time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][1].value, 0) if sheet[row][
#                 1].value else ""
#         except TypeError:
#             time_from_china = sheet[row][1].value if sheet[row][1].value else ""
#         total_cost = str(sheet[row][18].value) if sheet[row][18].value else ""
#         try:
#             if time_from_china >= datetime.datetime.strptime('21.11.2023', '%d.%m.%Y'):
#                 if sheet[row][15].value > 60:
#                     transportation_tariff_with_factor = float(
#                         sheet[row][15].value + self.factor_volume_20)
#                 else:
#                     transportation_tariff_with_factor = float(sheet[row][15].value + self.factor_kg_02)
#             else:
#                 transportation_tariff_with_factor = None
#         except TypeError:
#             transportation_tariff_with_factor = None
#         if transportation_tariff_with_factor:
#             if sheet[row][15].value > 60:
#                 total_cost_with_factor = float(total_cost) + self.factor_volume_20 * float(volume)
#             else:
#                 total_cost_with_factor = float(total_cost) + self.factor_kg_02 * float(weight)
#         else:
#             total_cost_with_factor = None
#         check = False
#         old_articles = CargoArticle.objects.filter(article=article)
#         for old in old_articles:
#             if old.article == article and old.name_goods == name_goods and old.status == 'В пути':
#                 check = True
#                 self.message['warning_articles'].append(
#                     f"Артикул '{old.article}' со статусом '{old.status}' и датой '{(old.time_from_china + datetime.timedelta(hours=3)).strftime('%d-%m-%Y')}' - уже существует")
#                 break
#         if not check:
#             CargoArticle.objects.create(
#                 article=article,
#                 carrier=carrier,
#                 name_goods=name_goods,
#                 number_of_seats=number_of_seats,
#                 weight=weight,
#                 volume=volume,
#                 transportation_tariff=transportation_tariff,
#                 transportation_tariff_with_factor=transportation_tariff_with_factor,
#                 cost_goods=cost_goods,
#                 insurance_cost=insurance_cost,
#                 packaging_cost=packaging_cost,
#                 time_from_china=make_aware(time_from_china),
#                 total_cost=total_cost,
#                 total_cost_with_factor=total_cost_with_factor,
#                 cargo_id=file_carrier,
#             )
#             self.message['success_articles'].append(article)
# elif file_carrier.name_carrier == 'Гелик':
# carrier = 'Гелик'
# if settings.DEBUG:
#     workbook = xlrd.open_workbook(f"{os.getcwd()}/media/{file_carrier.file_path}")
# else:
#     workbook = xlrd.open_workbook(f"{os.getcwd()}/leader_cargo/media/{file_carrier.file_path}")
# sheet = workbook.sheet_by_index(0)
# for row in range(7, sheet.nrows):
#     if sheet[row][1].value:
#         if sheet[2][1].value.find(":") != -1:
#             article = str(sheet[2][1].value.split(":")[1].replace(" ", ""))
#         else:
#             article = str(sheet[2][1].value.split("：")[1].replace(" ", ""))
#         name_goods = str(sheet[row][1].value) if sheet[row][1].value else ""
#         number_of_seats = str(sheet[row][2].value)
#         weight = str(sheet[row][3].value) if sheet[row][3].value else ""
#         volume = str(sheet[row][4].value) if sheet[row][4].value else ""
#         transportation_tariff = str(sheet[row][5].value) if sheet[row][5].value else ""
#         cost_goods = str(sheet[row][8].value) if sheet[row][8].value else ""
#         insurance_cost = str(sheet[row][9].value) if sheet[row][9].value else ""
#         packaging_cost = str(sheet[row][6].value) if sheet[row][6].value else ""
#         if sheet[4][7].value.find(":") != -1:
#             try:
#                 time_from_china = xlrd.xldate.xldate_as_datetime(
#                     sheet[4][7].value.split(":")[1].replace(" ", ""), 0)
#             except TypeError:
#                 try:
#                     time_from_china = datetime.datetime.strptime(
#                         sheet[4][7].value.split(":")[1].replace(" ", ""), '%Y/%m/%d')
#                 except:
#                     time_from_china = datetime.datetime.strptime(
#                         sheet[4][7].value.split(":")[1].replace(" ", ""), '%Y-%m-%d')
#         else:
#             try:
#                 time_from_china = xlrd.xldate.xldate_as_datetime(
#                     sheet[4][7].value.split("：")[1].replace(" ", ""), 0)
#             except TypeError:
#                 try:
#                     time_from_china = datetime.datetime.strptime(
#                         sheet[4][7].value.split("：")[1].replace(" ", ""), '%Y/%m/%d')
#                 except:
#                     time_from_china = datetime.datetime.strptime(
#                         sheet[4][7].value.split("：")[1].replace(" ", ""), '%Y-%m-%d')
#         total_cost = str(sheet[row][11].value) if sheet[row][11].value else ""
#         check = False
#         old_articles = CargoArticle.objects.filter(article=article)
#         for old in old_articles:
#             if old.article == article and old.name_goods == name_goods and old.status == 'В пути':
#                 check = True
#                 self.message['warning_articles'].append(
#                     f"Артикул '{old.article}' со статусом '{old.status}' и датой '{(old.time_from_china + datetime.timedelta(hours=3)).strftime('%d-%m-%Y')}' - уже существует")
#                 break
#         if not check:
#             CargoArticle.objects.create(
#                 article=article,
#                 carrier=carrier,
#                 name_goods=name_goods,
#                 number_of_seats=number_of_seats,
#                 weight=weight,
#                 volume=volume,
#                 transportation_tariff=transportation_tariff,
#                 cost_goods=cost_goods,
#                 insurance_cost=insurance_cost,
#                 packaging_cost=packaging_cost,
#                 time_from_china=make_aware(time_from_china),
#                 total_cost=total_cost,
#                 cargo_id=file_carrier,
#             )
#             self.message['success_articles'].append(article)
#     else:
#         break

class CarrierFilesView(MyLoginMixin, DataMixin, CreateView):
    model = CargoFiles
    form_class = AddCarrierFilesForm
    template_name = 'analytics/carrier_files.html'
    context_object_name = 'all_articles'
    login_url = reverse_lazy('main:login')
    role_have_perm = ['Супер Администратор', 'РОП', 'Логист', 'Менеджер']
    success_url = reverse_lazy('analytics:carrier')
    message = dict()
    message['update'] = False
    factor_kg_01 = 0.1
    factor_kg_02 = 0.2
    factor_volume_10 = 10
    factor_volume_20 = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_pk_string'] = str(self.request.user.pk)
        context['all_articles'] = CargoArticle.objects.all()
        if self.request.user.role == 'Логист' or self.request.user.role == 'Супер Администратор':
            context['count_empty_responsible_manager'] = context['all_articles'].filter(responsible_manager=None).count()
            if 100 - int(context['count_empty_responsible_manager'] / context['all_articles'].count() * 100) == 100 and context['count_empty_responsible_manager'] != 0:
                context['pb_count_empty_responsible_manager'] = 99
            else:
                context['pb_count_empty_responsible_manager'] = 100 - int(context['count_empty_responsible_manager'] / context['all_articles'].count() * 100)
            context['count_empty_path_format'] = context['all_articles'].filter(path_format=None).count()
            if 100 - int(context['count_empty_path_format'] / context['all_articles'].count() * 100) == 100 and context['count_empty_path_format'] != 0:
                context['pb_count_empty_path_format'] = 99
            else:
                context['pb_count_empty_path_format'] = 100 - int(context['count_empty_path_format'] / context['all_articles'].count() * 100)
            context['all_article_with_empty_responsible_manager'] = context['all_articles'].filter(responsible_manager=None).values('article')
            context['all_article_with_empty_path_format'] = context['all_articles'].filter(path_format=None).values('article')
            context['all_articles_without_insurance'] = context['all_articles'].filter(insurance_cost__in=[None, '']).filter(time_from_china__gte=make_aware(datetime.datetime.now() - datetime.timedelta(days=11))).values('article', 'weight',
                                                                                                                                                                                                                            'time_from_china')
            context['new_all_articles_without_insurance'] = []
            for art in context['all_articles_without_insurance']:
                if float(art['weight'].replace(" ", "").replace(",", ".")) > 10:
                    context['new_all_articles_without_insurance'].append(art)
            context['all_articles_without_insurance'] = context['new_all_articles_without_insurance']
            context['count_articles_without_insurance'] = len(context['all_articles_without_insurance'])

        context['count_notifications'] = context['all_articles'].filter(status='Прибыл в РФ').filter(time_cargo_release=None).count()
        context['all_articles_not_issued'] = context['all_articles'].filter(paid_by_the_client_status='Оплачено полностью').filter(time_cargo_release=None)
        context['count_all_articles_not_issued'] = context['all_articles'].filter(paid_by_the_client_status='Оплачено полностью').filter(time_cargo_release=None).count()

        if self.request.user.role == 'Менеджер':
            context['all_articles'] = context['all_articles'].filter(responsible_manager=f'{self.request.user.pk}')

        if self.request.user.role == 'РОП':
            managers_pk = [self.request.user.pk]
            context['managers'] = [{
                "pk": f"{self.request.user.pk}",
                "fi": f"{self.request.user.last_name} {self.request.user.first_name}"
            }]
            for user in CustomUser.objects.filter(role='Менеджер').filter(town=self.request.user.town).values('pk', 'last_name', 'first_name'):
                context['managers'].append({
                    "pk": f"{user['pk']}",
                    "fi": f"{user['last_name']} {user['first_name']}"
                })
                managers_pk.append(user['pk'])
            context['all_articles'] = context['all_articles'].filter(responsible_manager__in=managers_pk)
        elif self.request.user.role == 'Логист':
            context['managers'] = []
            for user in CustomUser.objects.filter(role__in=['Менеджер', 'РОП'], status=True).values('pk', 'last_name', 'first_name').order_by('last_name'):
                context['managers'].append({
                    "pk": f"{user['pk']}",
                    "fi": f"{user['last_name']} {user['first_name']}"
                })

        context['count_notifications'] = context['all_articles'].filter(status='Прибыл в РФ').filter(time_cargo_release=None).count()

        if self.request.htmx:
            if self.request.htmx.trigger == 'article_search':
                q = self.request.GET.get('q') if self.request.GET.get('q') is not None else ''
                context['q'] = q
                context['all_articles'] = context['all_articles'].filter(article__iregex=q)

        if self.request.GET.get('status') and self.request.GET.get('status') != 'Статус прибытия':
            context['status_now'] = self.request.GET.get('status')
            context['all_articles'] = context['all_articles'].filter(status=self.request.GET.get('status'))
        else:
            context['status_now'] = 'Статус прибытия'

        if self.request.GET.get('responsible_manager') and self.request.GET.get('responsible_manager') != 'Все менеджеры':
            context['responsible_manager_current'] = self.request.GET.get('responsible_manager')
            context['all_articles'] = context['all_articles'].filter(responsible_manager=context['responsible_manager_current'])
        else:
            context['responsible_manager_current'] = 'Все менеджеры'

        if self.request.GET.get('date'):
            context['date_current'] = self.request.GET.get('date')
            context['all_articles'] = context['all_articles'].filter(time_from_china__gte=make_aware(datetime.datetime.strptime(context['date_current'], '%Y-%m-%d')))
        else:
            if not self.request.htmx:
                if self.request.user.role == 'Логист':
                    first_day = (datetime.datetime.now() - datetime.timedelta(days=28)).replace(day=1)
                    context['date_current'] = first_day.strftime('%Y-%m-%d')
                    context['all_articles'] = context['all_articles'].filter(time_from_china__gte=make_aware(first_day))

        if self.request.GET.get('end_date'):
            context['end_date_current'] = self.request.GET.get('end_date')
            context['all_articles'] = context['all_articles'].filter(time_from_china__lte=make_aware(datetime.datetime.strptime(context['end_date_current'], '%Y-%m-%d')))

        if self.request.GET.get('paid_by_the_client') and self.request.GET.get('paid_by_the_client') != 'Оплата клиентом':
            context['paid_by_the_client_current'] = self.request.GET.get('paid_by_the_client')
            context['all_articles'] = context['all_articles'].filter(paid_by_the_client_status=context['paid_by_the_client_current'])
        else:
            context['paid_by_the_client_current'] = 'Оплата клиентом'

        if self.request.GET.get('paid_to_the_carrier') and self.request.GET.get('paid_to_the_carrier') != 'Оплата перевозчику':
            context['paid_to_the_carrier_current'] = self.request.GET.get('paid_to_the_carrier')
            context['all_articles'] = context['all_articles'].filter(payment_to_the_carrier_status=context['paid_to_the_carrier_current'])
        else:
            context['paid_to_the_carrier_current'] = 'Оплата перевозчику'

        if self.request.GET.get('carrier') and self.request.GET.get('carrier') != 'Все перевозчики':
            context['carrier_now'] = self.request.GET.get('carrier')
            context['all_articles'] = context['all_articles'].filter(carrier=self.request.GET.get('carrier'))
        else:
            context['carrier_now'] = 'Все перевозчики'

        context['form_article'] = []
        context['all_weight'] = 0
        context['all_volume'] = 0
        context['all_prr'] = 0
        context['all_tat'] = 0
        for article in context['all_articles']:
            if article.status:
                context['all_weight'] = context['all_weight'] + float(article.weight.replace(" ", "").replace(",", "."))
                context['all_volume'] = context['all_volume'] + float(article.volume.replace(" ", "").replace(",", "."))
                if article.prr:
                    context['all_prr'] = context['all_prr'] + float(article.prr.replace(" ", "").replace(",", "."))
                if article.tat_cost:
                    context['all_tat'] = context['all_tat'] + float(article.tat_cost.replace(" ", "").replace(",", "."))
        context['table_paginator'] = Paginator(context['all_articles'], 50)
        page_number = self.request.GET.get('page', 1)
        context['table_paginator_obj'] = context['table_paginator'].get_page(page_number)
        context['all_articles'] = context['table_paginator_obj']
        context['set_query'] = ''
        for req in self.request.GET:
            if req != 'page':
                context['set_query'] += f'&{req}={self.request.GET.get(req)}'
        context['vputi'] = 'В пути'
        context['pribil'] = 'Прибыл в РФ'
        if self.message['update']:
            context['message'] = self.message
        else:
            context['message'] = []
        self.message['update'] = False
        c_def = self.get_user_context(title="Учет грузов")
        return dict(list(context.items()) + list(c_def.items()))

    def get_template_names(self):
        if self.request.htmx:
            return "analytics/components/table_for_carrier_files.html"
        else:
            return 'analytics/carrier_files.html'

    def form_valid(self, form, **kwargs):
        file_carrier = form.save(commit=False)
        file_carrier.save()
        self.message['success_articles'] = []
        self.message['warning_articles'] = []
        self.message['info_articles'] = []
        self.message['error'] = []

        try:
            if file_carrier.name_carrier == 'Ян':
                carrier = 'Ян'
                if settings.DEBUG:
                    dataframe = openpyxl.load_workbook(os.path.join(str(os.getcwd()), 'media', str(file_carrier.file_path)), data_only=True)
                else:
                    dataframe = openpyxl.load_workbook(os.path.join(str(os.getcwd()), 'leader_cargo/media', str(file_carrier.file_path)), data_only=True)
                sheet = dataframe.get_sheet_by_name('运单')
                address_transportation_cost = ""
                for row in range(6, sheet.max_row):
                    if sheet[row][0].value == '送货费':
                        address_transportation_cost = float(sheet[row][13].value) / (row - 6)
                for row in range(6, sheet.max_row):
                    if sheet[row][0].value and sheet[row][0].value != '送货费':
                        article = str(sheet[row][0].value)
                        name_goods = str(sheet[row][1].value) if sheet[row][1].value else ""
                        number_of_seats = str(sheet[row][2].value)
                        weight = str(sheet[row][4].value) if sheet[row][4].value else ""
                        volume = str(sheet[row][5].value) if sheet[row][5].value else ""
                        transportation_tariff = str(sheet[row][6].value) if sheet[row][6].value else ""
                        cost_goods = str(sheet[row][7].value) if sheet[row][7].value else ""
                        insurance_cost = str(sheet[row][8].value) if sheet[row][8].value else ""
                        packaging_cost = str(sheet[row][9].value) if sheet[row][9].value else ""
                        if sheet[row][11].value:
                            time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][11].value, 0) if sheet[row][11].value else ""
                        else:
                            try:
                                time_from_china = xlrd.xldate.xldate_as_datetime(sheet[3][6].value, 0) if sheet[3][6].value else ""
                            except TypeError:
                                time_from_china = sheet[3][6].value
                        total_cost = str(sheet[row][13].value) if sheet[row][13].value else ""
                        try:
                            if time_from_china >= datetime.datetime.strptime('21.11.2023', '%d.%m.%Y'):
                                if sheet[row][6].value > 60:
                                    transportation_tariff_with_factor = float(sheet[row][6].value + self.factor_volume_10)
                                else:
                                    transportation_tariff_with_factor = float(sheet[row][6].value + self.factor_kg_01)
                            else:
                                transportation_tariff_with_factor = None
                        except TypeError:
                            transportation_tariff_with_factor = None
                        if transportation_tariff_with_factor:
                            if sheet[row][6].value > 60:
                                total_cost_with_factor = float(total_cost) + self.factor_volume_10 * float(volume)
                            else:
                                total_cost_with_factor = float(total_cost) + self.factor_kg_01 * float(weight)
                        else:
                            total_cost_with_factor = None
                        check = False
                        old_articles = CargoArticle.objects.filter(article=article)
                        for old in old_articles:
                            if old.article == article and old.name_goods == name_goods and old.weight == weight and old.time_from_china == make_aware(time_from_china):
                                check = True
                                old.carrier = carrier
                                old.number_of_seats = number_of_seats
                                old.volume = volume
                                old.transportation_tariff = transportation_tariff
                                if old.transportation_tariff_with_factor:
                                    old.transportation_tariff_with_factor = transportation_tariff_with_factor
                                    old.total_cost_with_factor = total_cost_with_factor
                                old.cost_goods = cost_goods
                                old.insurance_cost = insurance_cost
                                old.packaging_cost = packaging_cost
                                old.total_cost = total_cost
                                old.cargo_id = file_carrier
                                old.address_transportation_cost = address_transportation_cost
                                old.save()
                                self.message['info_articles'].append(f"Артикул '{old.article}' со статусом '{old.status}' "
                                                                     f"и датой '{(old.time_from_china + datetime.timedelta(hours=3)).strftime('%d-%m-%Y')}'")
                                break
                        if not check:
                            CargoArticle.objects.create(
                                article=article,
                                carrier=carrier,
                                name_goods=name_goods,
                                number_of_seats=number_of_seats,
                                weight=weight,
                                volume=volume,
                                transportation_tariff=transportation_tariff,
                                transportation_tariff_with_factor=transportation_tariff_with_factor,
                                cost_goods=cost_goods,
                                insurance_cost=insurance_cost,
                                packaging_cost=packaging_cost,
                                time_from_china=make_aware(time_from_china),
                                total_cost=total_cost,
                                total_cost_with_factor=total_cost_with_factor,
                                cargo_id=file_carrier,
                                address_transportation_cost=address_transportation_cost,
                            )
                            self.message['success_articles'].append(article)
                    else:
                        break
            elif file_carrier.name_carrier == 'Ян (полная машина)':
                carrier = 'Ян'
                if settings.DEBUG:
                    dataframe = openpyxl.load_workbook(os.path.join(str(os.getcwd()), 'media', str(file_carrier.file_path)), data_only=True)
                else:
                    dataframe = openpyxl.load_workbook(os.path.join(str(os.getcwd()), 'leader_cargo/media', str(file_carrier.file_path)), data_only=True)
                sheet = dataframe.active
                article = str(sheet[3][3].value)
                number_of_seats = str(sheet[3][4].value)
                weight = str(sheet[3][5].value) if sheet[3][5].value else ""
                volume = str(sheet[3][6].value) if sheet[3][6].value else ""
                address_transportation_cost = float(sheet[3][9].value)
                try:
                    time_from_china = xlrd.xldate.xldate_as_datetime(sheet[3][0].value, 0) if sheet[3][0].value else ""
                except TypeError:
                    time_from_china = sheet[3][0].value
                try:
                    make_aware(time_from_china)
                except:
                    time_from_china = datetime.datetime.strptime(time_from_china, '%Y.%m.%d')
                total_cost = str(sheet[3][10].value) if sheet[3][10].value else ""
                check = False
                old_articles = CargoArticle.objects.filter(article=article)
                for old in old_articles:
                    if old.article == article and old.status == 'В пути':
                        check = True
                        self.message['warning_articles'].append(f"Артикул '{old.article}' со статусом '{old.status}' и датой '{(old.time_from_china + datetime.timedelta(hours=3)).strftime('%d-%m-%Y')}' - уже существует")
                        break
                if not check:
                    CargoArticle.objects.create(
                        article=article,
                        carrier=carrier,
                        number_of_seats=number_of_seats,
                        weight=weight,
                        volume=volume,
                        time_from_china=make_aware(time_from_china),
                        total_cost=total_cost,
                        cargo_id=file_carrier,
                        address_transportation_cost=address_transportation_cost,
                    )
                    self.message['success_articles'].append(article)
            elif file_carrier.name_carrier == 'Валька':
                carrier = 'Валька'
                if settings.DEBUG:
                    dataframe = openpyxl.load_workbook(f"{os.getcwd()}/media/{file_carrier.file_path}", data_only=True)
                else:
                    dataframe = openpyxl.load_workbook(f"{os.getcwd()}/leader_cargo/media/{file_carrier.file_path}", data_only=True)
                sheet = dataframe.active
                for row in range(790, sheet.max_row):
                    if sheet[row][4].value and sheet[row][3].value and sheet[row][2].value:
                        article = str(sheet[row][4].value)
                        name_goods = str(sheet[row][6].value) if sheet[row][6].value else ""
                        number_of_seats = str(sheet[row][7].value)
                        weight = str(sheet[row][8].value) if sheet[row][8].value else ""
                        volume = str(sheet[row][9].value) if sheet[row][9].value else ""
                        transportation_tariff = str(round(float(sheet[row][12].value) / float(sheet[row][8].value), 2))
                        cost_goods = str(sheet[row][13].value) if sheet[row][13].value else ""
                        insurance_cost = str(sheet[row][14].value) if sheet[row][14].value else ""
                        packaging_cost = str(sheet[row][15].value) if sheet[row][15].value else ""
                        try:
                            time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][3].value, 0) if sheet[row][3].value else ""
                        except TypeError:
                            time_from_china = sheet[row][3].value
                        total_cost = str(sheet[row][16].value) if sheet[row][16].value else ""
                        try:
                            if time_from_china >= datetime.datetime.strptime('21.11.2023', '%d.%m.%Y'):
                                if round(float(sheet[row][12].value) / float(sheet[row][8].value), 2) > 60:
                                    transportation_tariff_with_factor = float(round(float(sheet[row][12].value) / float(sheet[row][8].value), 2) + self.factor_volume_20)
                                else:
                                    transportation_tariff_with_factor = float(round(float(sheet[row][12].value) / float(sheet[row][8].value), 2) + self.factor_kg_02)
                            else:
                                transportation_tariff_with_factor = None
                        except TypeError:
                            transportation_tariff_with_factor = None
                        if transportation_tariff_with_factor:
                            if round(float(sheet[row][12].value) / float(sheet[row][8].value), 2) > 60:
                                total_cost_with_factor = float(total_cost) + self.factor_volume_20 * float(volume)
                            else:
                                total_cost_with_factor = float(total_cost) + self.factor_kg_02 * float(weight)
                        else:
                            total_cost_with_factor = None
                        check = False
                        old_articles = CargoArticle.objects.filter(article=article)
                        for old in old_articles:
                            if old.article == article and old.name_goods == name_goods and old.number_of_seats == number_of_seats and old.cost_goods == cost_goods and old.insurance_cost == insurance_cost \
                                    and old.weight == weight and old.volume == volume and old.transportation_tariff == transportation_tariff and old.packaging_cost == packaging_cost \
                                    and old.time_from_china == make_aware(time_from_china) and old.total_cost == total_cost:
                                check = True
                                break
                        if not check:
                            CargoArticle.objects.create(
                                article=article,
                                carrier=carrier,
                                name_goods=name_goods,
                                number_of_seats=number_of_seats,
                                weight=weight,
                                volume=volume,
                                transportation_tariff=transportation_tariff,
                                transportation_tariff_with_factor=transportation_tariff_with_factor,
                                cost_goods=cost_goods,
                                insurance_cost=insurance_cost,
                                packaging_cost=packaging_cost,
                                time_from_china=make_aware(time_from_china),
                                total_cost=total_cost,
                                total_cost_with_factor=total_cost_with_factor,
                                cargo_id=file_carrier,
                            )
                            self.message['success_articles'].append(article)
                    else:
                        break
            elif file_carrier.name_carrier == 'Мурад':
                carrier = 'Мурад'
                if settings.DEBUG:
                    dataframe = openpyxl.load_workbook(f"{os.getcwd()}/media/{file_carrier.file_path}", data_only=True)
                else:
                    dataframe = openpyxl.load_workbook(f"{os.getcwd()}/leader_cargo/media/{file_carrier.file_path}", data_only=True)
                sheet = dataframe.active
                for row in range(2, sheet.max_row + 1):
                    if sheet[row][5].value:
                        if sheet[row][5].value.find("（") != -1:
                            article = str(sheet[row][5].value.split("（")[0].replace(" ", "")) + '\n' + str(sheet[row][3].value)
                            if sheet[row][5].value.find("）") != -1:
                                name_goods = str(sheet[row][5].value.split("（")[1].replace(" ", "").replace("）", ""))
                            else:
                                name_goods = str(sheet[row][5].value.split("（")[1].replace(" ", "").replace(")", ""))
                        else:
                            article = str(sheet[row][5].value.split("(")[0].replace(" ", "")) + '\n' + str(sheet[row][3].value)
                            if sheet[row][5].value.find("）") != -1:
                                name_goods = str(sheet[row][5].value.split("(")[1].replace(" ", "").replace("）", ""))
                            else:
                                name_goods = str(sheet[row][5].value.split("(")[1].replace(" ", "").replace(")", ""))
                        number_of_seats = str(sheet[row][4].value)
                        weight = str(sheet[row][6].value) if sheet[row][4].value else ""
                        volume = str(sheet[row][7].value) if sheet[row][5].value else ""
                        transportation_tariff = str(sheet[row][15].value) if sheet[row][15].value else ""
                        cost_goods = str(sheet[row][12].value) if sheet[row][12].value else ""
                        insurance_cost = str(sheet[row][13].value) if sheet[row][13].value else ""
                        packaging_cost = str(sheet[row][14].value) if sheet[row][14].value else ""
                        try:
                            time_from_china = xlrd.xldate.xldate_as_datetime(sheet[row][1].value, 0) if sheet[row][1].value else ""
                        except TypeError:
                            time_from_china = sheet[row][1].value if sheet[row][1].value else ""
                        total_cost = str(sheet[row][18].value) if sheet[row][18].value else ""
                        try:
                            if time_from_china >= datetime.datetime.strptime('21.11.2023', '%d.%m.%Y'):
                                if sheet[row][15].value > 60:
                                    transportation_tariff_with_factor = float(sheet[row][15].value + self.factor_volume_20)
                                else:
                                    transportation_tariff_with_factor = float(sheet[row][15].value + self.factor_kg_02)
                            else:
                                transportation_tariff_with_factor = None
                        except TypeError:
                            transportation_tariff_with_factor = None
                        if transportation_tariff_with_factor:
                            if sheet[row][15].value > 60:
                                total_cost_with_factor = float(total_cost) + self.factor_volume_20 * float(volume)
                            else:
                                total_cost_with_factor = float(total_cost) + self.factor_kg_02 * float(weight)
                        else:
                            total_cost_with_factor = None
                        check = False
                        old_articles = CargoArticle.objects.filter(article=article)
                        for old in old_articles:
                            if old.article == article and old.name_goods == name_goods and old.status == 'В пути':
                                check = True
                                self.message['warning_articles'].append(f"Артикул '{old.article}' со статусом '{old.status}' и датой '{(old.time_from_china + datetime.timedelta(hours=3)).strftime('%d-%m-%Y')}' - уже существует")
                                break
                        if not check:
                            CargoArticle.objects.create(
                                article=article,
                                carrier=carrier,
                                name_goods=name_goods,
                                number_of_seats=number_of_seats,
                                weight=weight,
                                volume=volume,
                                transportation_tariff=transportation_tariff,
                                transportation_tariff_with_factor=transportation_tariff_with_factor,
                                cost_goods=cost_goods,
                                insurance_cost=insurance_cost,
                                packaging_cost=packaging_cost,
                                time_from_china=make_aware(time_from_china),
                                total_cost=total_cost,
                                total_cost_with_factor=total_cost_with_factor,
                                cargo_id=file_carrier,
                            )
                            self.message['success_articles'].append(article)
            elif file_carrier.name_carrier == 'Гелик':
                carrier = 'Гелик'
                if settings.DEBUG:
                    workbook = xlrd.open_workbook(f"{os.getcwd()}/media/{file_carrier.file_path}")
                else:
                    workbook = xlrd.open_workbook(f"{os.getcwd()}/leader_cargo/media/{file_carrier.file_path}")
                sheet = workbook.sheet_by_index(0)
                for row in range(7, sheet.nrows):
                    if sheet[row][1].value:
                        if sheet[2][1].value.find(":") != -1:
                            article = str(sheet[2][1].value.split(":")[1].replace(" ", ""))
                        else:
                            article = str(sheet[2][1].value.split("：")[1].replace(" ", ""))
                        name_goods = str(sheet[row][1].value) if sheet[row][1].value else ""
                        number_of_seats = str(sheet[row][2].value)
                        weight = str(sheet[row][3].value) if sheet[row][3].value else ""
                        volume = str(sheet[row][4].value) if sheet[row][4].value else ""
                        transportation_tariff = str(sheet[row][5].value) if sheet[row][5].value else ""
                        cost_goods = str(sheet[row][8].value) if sheet[row][8].value else ""
                        insurance_cost = str(sheet[row][9].value) if sheet[row][9].value else ""
                        packaging_cost = str(sheet[row][6].value) if sheet[row][6].value else ""
                        if sheet[4][7].value.find(":") != -1:
                            try:
                                time_from_china = xlrd.xldate.xldate_as_datetime(sheet[4][7].value.split(":")[1].replace(" ", ""), 0)
                            except TypeError:
                                time_from_china = datetime.datetime.strptime(sheet[4][7].value.split(":")[1].replace(" ", ""), '%Y/%m/%d')
                        else:
                            try:
                                time_from_china = xlrd.xldate.xldate_as_datetime(sheet[4][7].value.split("：")[1].replace(" ", ""), 0)
                            except TypeError:
                                time_from_china = datetime.datetime.strptime(sheet[4][7].value.split("：")[1].replace(" ", ""), '%Y/%m/%d')
                        total_cost = str(sheet[row][11].value) if sheet[row][11].value else ""
                        check = False
                        old_articles = CargoArticle.objects.filter(article=article)
                        for old in old_articles:
                            if old.article == article and old.name_goods == name_goods and old.status == 'В пути':
                                check = True
                                self.message['warning_articles'].append(f"Артикул '{old.article}' со статусом '{old.status}' и датой '{(old.time_from_china + datetime.timedelta(hours=3)).strftime('%d-%m-%Y')}' - уже существует")
                                break
                        if not check:
                            CargoArticle.objects.create(
                                article=article,
                                carrier=carrier,
                                name_goods=name_goods,
                                number_of_seats=number_of_seats,
                                weight=weight,
                                volume=volume,
                                transportation_tariff=transportation_tariff,
                                cost_goods=cost_goods,
                                insurance_cost=insurance_cost,
                                packaging_cost=packaging_cost,
                                time_from_china=make_aware(time_from_china),
                                total_cost=total_cost,
                                cargo_id=file_carrier,
                            )
                            self.message['success_articles'].append(article)
                    else:
                        break
        except Exception as e:
            self.message['error'].append(e)
        self.message['update'] = True
        return redirect('{}?{}'.format(reverse('analytics:carrier'), self.request.META['QUERY_STRING']))


def change_article_status(request, article_id):
    role_have_perm = ['Супер Администратор', 'Логист']
    if request.user.role in role_have_perm:
        article = CargoArticle.objects.get(pk=article_id)
        article.update_status_article()
    return redirect(request.META.get('HTTP_REFERER') + f'#article-{article_id}')


class EditTableArticleView(MyLoginMixin, DataMixin, UpdateView):
    model = CargoArticle
    form_class = EditTableArticleForm
    pk_url_kwarg = 'article_id'
    role_have_perm = ['Супер Администратор', 'Логист']
    success_url = reverse_lazy('analytics:carrier')
    login_url = reverse_lazy('main:login')

    def form_valid(self, form):
        file_carrier = form.save(commit=False)
        if form.cleaned_data['time_cargo_arrival_to_RF']:
            file_carrier.status = 'Прибыл в РФ'
        if form.cleaned_data['time_cargo_release']:
            file_carrier.status = 'Выдан'
        file_carrier.save()
        return redirect(self.request.META.get('HTTP_REFERER') + f'#article-{self.object.pk}')


class EditTransportTariff(MyLoginMixin, DataMixin, UpdateView):
    model = CargoArticle
    form_class = EditTransportationTariffForClients
    pk_url_kwarg = 'article_id'
    context_object_name = 'article'
    role_have_perm = ['Супер Администратор', 'РОП', 'Менеджер']
    success_url = reverse_lazy('analytics:carrier')
    login_url = reverse_lazy('main:login')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Изменение транспортного тарифа клиента")
        return dict(list(context.items()) + list(c_def.items()))

    def get_template_names(self):
        if self.request.htmx.target == 'hx-modal-main':
            return 'analytics/logistic_main/modal/modalEditClientsTariff.html'
        else:
            return 'analytics/logistic_main/modal/modalEditClientsTariff.html'

    def form_valid(self, form):
        article = form.save(commit=False)
        article.save()
        return redirect(self.request.META.get('HTTP_REFERER') + f'#article-{self.object.pk}')


def amountOfFund(request):
    months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
    years = [2023, 2024, 2025, 2026, 2027, 2028]
    now = datetime.datetime.now()
    month = request.GET.get('month', months[now.month - 1])
    year = request.GET.get('year', now.year)
    articles = CargoArticle.objects.filter(total_cost_with_factor__isnull=False)
    if request.htmx:
        if request.htmx.target == 'amountOfFundContent':
            if month != 'Все месяца':
                for index, m in enumerate(months):
                    if m == month:
                        month = index + 1
                        break
                articles = articles.filter(time_from_china__month=month)
            if year != 'Все года':
                articles = articles.filter(time_from_china__year=year)
            amount_of_fund = 0
            count_articles_with_factor = articles.count()
            for art in articles:
                amount_of_fund = amount_of_fund + (art.total_cost_with_factor - float(art.total_cost.replace(" ", "").replace(",", ".")))
            return render(request, 'analytics/components/modal/htmxAmountFundResults.html', {'amount_of_fund': amount_of_fund,
                                                                                             'count_articles_with_factor': count_articles_with_factor,
                                                                                             'month': month,
                                                                                             'year': year,
                                                                                             'months': months,
                                                                                             'years': years})
    amount_of_fund = 0
    month_number = 0
    for index, m in enumerate(months):
        if m == month:
            month_number = index + 1
            break
    articles = articles.filter(time_from_china__month=month_number, time_from_china__year=year)
    count_articles_with_factor = articles.count()
    for art in articles:
        amount_of_fund = amount_of_fund + (art.total_cost_with_factor - float(art.total_cost.replace(" ", "").replace(",", ".")))
    return render(request, 'analytics/components/modal/htmxModalAmountFund.html', {'amount_of_fund': amount_of_fund,
                                                                                   'count_articles_with_factor': count_articles_with_factor,
                                                                                   'month': month,
                                                                                   'year': year,
                                                                                   'months': months,
                                                                                   'years': years})


def edit_article_in_table(request, article_id):
    article = CargoArticle.objects.get(pk=article_id)
    form_article = EditTableArticleForm(instance=article,
                                        initial={
                                            'time_cargo_arrival_to_RF': (article.time_cargo_arrival_to_RF + datetime.timedelta(hours=3)).strftime("%Y-%m-%d")
                                            if article.time_cargo_arrival_to_RF else article.time_cargo_arrival_to_RF,
                                            'time_cargo_release': (article.time_cargo_release + datetime.timedelta(hours=3)).strftime("%Y-%m-%d")
                                            if article.time_cargo_release else article.time_cargo_release,
                                        })
    return render(request, 'analytics/components/modal/htmxModalEditArticleForLogist.html', {'article': article, 'form_article': form_article})


def delete_article_in_table(request, article_id):
    article = CargoArticle.objects.get(pk=article_id)
    return render(request, 'analytics/components/modal/htmxModalDeleteArticleForLogist.html', {'article': article})


def paid_by_the_client_status(request, article_id):
    role_have_perm = ['Супер Администратор', 'Менеджер', 'РОП', 'Логист']
    if request.user.role in role_have_perm:
        article = CargoArticle.objects.get(pk=article_id)
        if request.user.pk == int(article.responsible_manager.pk):
            if request.POST:
                if article.file_payment_by_client.all():
                    article.paid_by_the_client_status = request.POST.get('paid_by_the_client_status')
                    article.time_paid_by_the_client_status = make_aware(datetime.datetime.now())
                    article.save()
                    return render(request, 'analytics/logistic_main/edit_Td_PaidByTheClientStatus.html',
                                  {'article': article})
            if article.file_payment_by_client.all() and article.paid_by_the_client_status == 'Оплачено полностью':
                return render(request, 'analytics/logistic_main/modal/modalViewArticlePaidByClientStatus.html',
                              {'article': article})
            form = EditPaidByTheClientArticleForm(instance=article)
            message = False
            if request.htmx.target == 'modal-alert-message':
                message = True
                if article.file_payment_by_client.all():
                    message = False
                return render(request, 'analytics/logistic_main/partial/modal_error_message.html',
                              {'message': message})

            return render(request, 'analytics/logistic_main/modal/modalEditArticlePaidByClientStatus.html',
                          {'article': article,
                           'form': form,
                           'message': message})
        else:
            return render(request, 'analytics/logistic_main/modal/modalViewArticlePaidByClientStatus.html',
                          {'article': article})
    else:
        return redirect('analytics:carrier')


def add_payment_file(request, article_id):
    role_have_perm = ['Супер Администратор', 'Менеджер', 'РОП']
    article = CargoArticle.objects.get(pk=article_id)
    form = EditPaidByTheClientArticleForm(instance=article)
    if request.user.role in role_have_perm:
        for f in request.FILES.getlist('files'):
            article.file_payment_by_client.create(name=f, file_path=f)
        return render(request,
                      'analytics/logistic_main/create_payment_file.html',
                      {'article': article,
                       'form': form})
    else:
        return render(request,
                      'analytics/logistic_main/create_payment_file.html',
                      {'article': article,
                       'form': form})


def delete_payment_file(request, article_id, file_id):
    role_have_perm = ['Супер Администратор', 'Менеджер', 'РОП']
    article = CargoArticle.objects.get(pk=article_id)
    file = PaymentDocumentsForArticles.objects.get(pk=file_id)
    form = EditPaidByTheClientArticleForm(instance=article)
    if request.user.role in role_have_perm:
        file.delete()
        return render(request,
                      'analytics/logistic_main/create_payment_file.html',
                      {'article': article,
                       'form': form})
    else:
        return render(request,
                      'analytics/logistic_main/create_payment_file.html',
                      {'article': article,
                       'form': form})


def change_article_for_manager(request, article_id):
    role_have_perm = ['Супер Администратор', 'Менеджер', 'РОП']
    if request.user.role in role_have_perm:
        article = CargoArticle.objects.get(pk=article_id)
        article.paid_by_the_client_status = request.GET.get('paid_by_the_client_status')
        article.save()
        return redirect(request.META.get('HTTP_REFERER') + f'#article-{article.pk}')
    return redirect(request.META.get('HTTP_REFERER'))


class DeleteArticleView(MyLoginMixin, DataMixin, DeleteView):
    model = CargoArticle
    pk_url_kwarg = 'article_id'
    role_have_perm = ['Супер Администратор', 'Логист']
    success_url = reverse_lazy('analytics:carrier')
    login_url = reverse_lazy('main:login')

    def delete(self, request, *args, **kwargs):
        article = self.get_object()
        article.delete()
        return redirect(request.META.get('HTTP_REFERER'))

@method_decorator(csrf_exempt, name='dispatch')
class UpdateOrderingView(View):
    def post(self, request, *args, **kwargs):
        try:
            # Получаем новый порядок (используем 'ordering', а не 'ordering[]')
            ordering = request.POST.getlist('ordering')

            # Удаляем 'undefined', если он есть
            ordering = [item for item in ordering if item != 'undefined']

            # Обновляем порядок в базе данных
            for index, road_id in enumerate(ordering):
                road = RoadsList.objects.get(pk=road_id)
                road.ordering = index + 1  # Новый индекс (начинается с 1)
                road.save()

            # Возвращаем успешный ответ
            return JsonResponse({'success': True})
        except Exception as e:
            # Возвращаем ошибку в случае исключения
            return JsonResponse({'success': False, 'error': str(e)}, status=400)